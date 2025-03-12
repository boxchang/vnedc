from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils import timezone
from django.utils.translation import get_language
from django.utils.translation import gettext_lazy as _
from openpyxl.utils import get_column_letter
from VNEDC.database import vnedc_database
from collection.forms import DailyInfoForm, ExcelUploadForm
from collection.models import ParameterDefine, Process_Type, Plant, Machine, Daily_Prod_Info, ParameterValue, \
    Daily_Prod_Info_Head, Lab_Parameter_Control, Parameter_Type
from jobs.database import scada_database
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from django.shortcuts import render
import re
import requests
import os
import hashlib
import base64
import pandas as pd

def create_vnedc_connection():
    pass

@login_required
def index(request):
    plants = Plant.objects.all()
    machs = Machine.objects.none()
    if request.method == 'POST':
        sData_date = request.POST.get('data_date')
        sPlant = request.POST.get('plant')
    else:
        sPlant = "GDNBR"  # Default is GDNBR
        sData_date = datetime.today()
        sData_date = sData_date.strftime("%Y-%m-%d")

    machs = Machine.objects.filter(plant=sPlant)
    parameter_list = []
    for tmp_mach in machs:
        daily_prod_info_heads = Daily_Prod_Info_Head.objects.filter(data_date=sData_date,
                                                                    mach=tmp_mach.mach_code).order_by('mach_id', 'line')
        tmp_mach.daily_prod_info_heads = daily_prod_info_heads

        defines = ParameterDefine.objects.filter(mach=tmp_mach.mach_code, auto_value=0)

        goal_count = 0
        for define in defines:
            if define.sampling_frequency == "6H" or define.sampling_frequency == None:
                goal_count += 4
            elif define.sampling_frequency == "12H":
                goal_count += 2

            # Setup Define Parameter
            # tmp_param = define.process_type.process_code + "_" + define.parameter_name
            # if not tmp_param in parameter_list:
            #    parameter_list.append(define.process_type.process_code + "_" + define.parameter_name)

        reach_count = 0

        tmp_msg = ""

        sql = f"""            
        select d.plant_id,d.mach_id,d.parameter_name,v.parameter_value,d.process_type_id from (SELECT *
          FROM [collection_parameterdefine] 
          where plant_id='{tmp_mach.plant}' and mach_id='{tmp_mach.mach_code}' and auto_value=0) d
          join 
          (select * from [collection_parametervalue] 
          where data_date = '{sData_date}' and plant_id='{tmp_mach.plant}' and mach_id='{tmp_mach.mach_code}' and parameter_value>0) v 
          on d.plant_id = v.plant_id and d.mach_id = v.mach_id and d.parameter_name = v.parameter_name 
          and d.process_type_id = v.process_type
            order by process_type_id
        """
        db = vnedc_database()
        results = db.select_sql_dict(sql)
        for result in results:
            if result["parameter_value"] != None and result["parameter_value"] > 0:
                reach_count += 1

                # Remove Completed Parameter
                # tmp_param = result["process_type_id"] + "_" + result["parameter_name"]
                # if tmp_param in parameter_list:
                #    parameter_list.remove(tmp_param)

        hit_rate_msg = f"{reach_count}/{goal_count}"

        hit_rate = 0
        if reach_count > 0:
            hit_rate = int(round(reach_count / goal_count, 2) * 100)
        tmp_mach.hit_rate_msg = hit_rate_msg
        tmp_mach.hit_rate = hit_rate

        # tmp_msg = "\r\n".join(parameter_list)
        # tmp_mach.msg = tmp_msg

    return render(request, 'collection/index.html', locals())


@login_required
def prod_info_reset(request):
    if request.method == 'POST':
        if 'plant' in request.session:
            del request.session['plant']

        if 'mach' in request.session:
            del request.session['mach']

        if 'data_date' in request.session:
            del request.session['data_date']
    return redirect(reverse('daily_info_create'))


@login_required
def page_init(request):
    plant = ""
    mach = ""
    data_date = ""

    if 'plant' in request.session:
        plant = request.session['plant']
    else:
        plant = request.POST.get('plant')

    if 'mach' in request.session:
        mach = request.session['mach']
    else:
        mach = request.POST.get('mach')

    if 'data_date' in request.session:
        data_date = request.session['data_date']
    else:
        data_date = request.POST.get('data_date')

    lang = get_language()

    return plant, mach, data_date, lang


@login_required
def prod_info_save(request):
    if request.method == 'POST':
        request.session['plant'] = request.POST.get('plant')
        request.session['mach'] = request.POST.get('mach')
        request.session['data_date'] = request.POST.get('data_date')
    return redirect(reverse('daily_info_create'))


@login_required
def record(request, process_code):
    sPlant, sMach, sData_date, lang = page_init(request)
    save = 0
    process_type = Process_Type.objects.filter(process_code=process_code).first()
    processes = Process_Type.objects.all().order_by('show_order')
    data_times = ['00', '06', '12', '18']
    plants = Plant.objects.all()

    if sPlant:
        machs = Machine.objects.filter(plant=sPlant)
    else:
        machs = None

    sql_01 = f"""
                SELECT product
                FROM [VNEDC].[dbo].[collection_daily_prod_info_head]
                where data_date = '{sData_date}' and mach_id = '{sMach}'
                """
    db = vnedc_database()
    results = db.select_sql_dict(sql_01)
    result = '/'.join(item['product'] for item in results)
    item_no = re.findall(r'\d+', result)[0] if len(result) > 0 else 0
    sql_02 = f"""
                select parameter_name, control_range_low as low_limit, control_range_high as high_limit
                from [VNEDC].[dbo].[collection_lab_parameter_control] 
                where item_no = {item_no} and mach_id = '{sMach}' and process_type = '{process_code}'
                """
    results = db.select_sql_dict(sql_02)
    limit = [[limit['parameter_name'], limit['low_limit'], limit['high_limit']] for limit in results]

    info = Daily_Prod_Info.objects.filter(plant=sPlant, mach=sMach, data_date=sData_date).first()
    alert_fields = []
    plant = Plant.objects.get(plant_code=sPlant)
    mach = Machine.objects.get(mach_code=sMach)
    if 'save_click_count' not in request.session:
        request.session['save_click_count'] = 0

    if request.method == 'POST':
        save = 1
        request.session['save_click_count'] += 1
        # request.session.modified = True

        if process_type:
            defines = ParameterDefine.objects.filter(plant=sPlant, mach=sMach, process_type=process_type)
            for define in defines:
                for time in data_times:
                    value = request.POST.get(define.parameter_name + '_' + time)
                    try:
                        if value:
                            ParameterValue.objects.update_or_create(plant=plant, mach=mach,
                                                                    data_date=sData_date,
                                                                    process_type=process_type.process_code,
                                                                    data_time=time,
                                                                    parameter_name=define.parameter_name,
                                                                    defaults={'parameter_value': value,
                                                                              'create_by': request.user,
                                                                              'update_by': request.user})

                            msg = _("Update Done")
                    except Exception as e:
                        print(e)
        # if request.session.modified == True:
        #     m_sql = f"""
        #                 SELECT parameter_name, parameter_value, data_time FROM [VNEDC].[dbo].[collection_parametervalue]
        #                 WHERE plant_id = '{sPlant}' AND mach_id = '{sMach}' AND data_date = '{sData_date}'
        #                 AND process_type = '{process_code}'
        #                 AND data_time = (
        #                     SELECT TOP 1 data_time FROM [VNEDC].[dbo].[collection_parametervalue]
        #                     WHERE plant_id = '{sPlant}' AND mach_id = '{sMach}' AND data_date = '{sData_date}'
        #                     AND process_type = '{process_code}' ORDER BY data_time DESC);
        #                 """
        #     values = db.select_sql_dict(m_sql)
        #     values_list = [[value['parameter_name'], value['parameter_value'], value['data_time']] for value in values]
        #     for limit_item in limit:
        #         for value_item in values_list:
        #             if str(limit_item[0]) in str(value_item[0]):
        #                 if (float(value_item[1]) < float(limit_item[1])) or (
        #                         float(value_item[1]) > float(limit_item[2])):
        #                     alert_fields.append(f"{value_item[0]}: {value_item[1]} ({str(limit_item[1])}-{str(limit_item[2])})")
        #     alert_fields.append(values_list[0][-1])
        #     if len(alert_fields) > 1:
        #         record = '\n\t+'.join(alert_fields[:-1])
        #         record = '\t+' + record
        #         message = f"Vui lòng kiểm tra lại giá trị của {process_code}:\n" + str(record) + f"\nVào lúc {alert_fields[-1]}:00"
        #         print(message)
        #         record_message(message)
        return redirect(reverse('record', kwargs={'process_code': process_code}))
    if process_type:
        defines = ParameterDefine.objects.filter(plant=sPlant, mach=sMach, process_type=process_type)
        for define in defines:
            values = ParameterValue.objects.filter(plant=sPlant, mach=sMach, data_date=sData_date,
                                                   process_type=process_type.process_code,
                                                   parameter_name=define.parameter_name).order_by('-create_at')

            if values:
                for time in data_times:
                    item = values.filter(data_time=time).first()
                    if item:
                        value = item.parameter_value
                        setattr(define, "T" + time, value)

    daily_prod_info_heads = Daily_Prod_Info_Head.objects.filter(data_date=sData_date).order_by('line')

    return render(request, 'collection/record.html', locals())


def record_message(msg):
    path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    wecom_file = os.path.join(path, 'Jobs', '../static/wecom/wecom_key.config')
    url = ''  # Add Wecom GD_MES group key
    if os.path.exists(wecom_file):
        with open(wecom_file, 'r') as file:
            url = file.read().strip()

    headers = {'Content-Type': 'application/json; charset=utf-8'}
    data = {
        "msgtype": "text",
        "text": {
            "content": '',
            # "mentioned_list": ["@all"],
        }
    }
    data["text"]["content"] = msg
    r = requests.post(url, headers=headers, json=data)
    return r.json()


def get_production_choices(end_date):
    date_obj = datetime.strptime(end_date, "%Y-%m-%d")
    date_obj_minus_one = date_obj - timedelta(days=365)
    start_date = date_obj_minus_one.strftime("%Y-%m-%d")

    sql1 = f"""
        SELECT distinct ProductItem as prod_name
        FROM [PMG_MES_WorkOrder] where SAP_FactoryDescr like '%NBR%' and WorkOrderDate between '{start_date}' and '{end_date}'
		order by ProductItem"""
    mes_db = mes_database()
    mes_rows = mes_db.select_sql_dict(sql1)
    mes_list = [mes_rows['prod_name'] for mes_rows in mes_rows]

    sql2 = f"""
        SELECT DISTINCT prod_name_a1 AS prod_name
        FROM [VNEDC].[dbo].[collection_daily_prod_info]
        WHERE data_date BETWEEN '{start_date}' AND '{end_date}'
        UNION
        SELECT DISTINCT prod_name_a2 AS prod_name
        FROM [VNEDC].[dbo].[collection_daily_prod_info]
        WHERE data_date BETWEEN '{start_date}' AND '{end_date}'
        UNION
        SELECT DISTINCT prod_name_b1 AS prod_name
        FROM [VNEDC].[dbo].[collection_daily_prod_info]
        WHERE data_date BETWEEN '{start_date}' AND '{end_date}'
        UNION
        SELECT DISTINCT prod_name_b2 AS prod_name
        FROM [VNEDC].[dbo].[collection_daily_prod_info]
        WHERE data_date BETWEEN '{start_date}' AND '{end_date}'
    """
    vnedc_db = vnedc_database()
    vnedc_rows = vnedc_db.select_sql_dict(sql2)
    vnedc_list = [vnedc_row['prod_name'] for vnedc_row in vnedc_rows]
    rows = list(set(mes_list + vnedc_list))
    choices = [('', '---')] + [(row, row) for row in rows]
    return choices

@login_required
def daily_info_create(request):
    sPlant, sMach, sData_date, lang = page_init(request)
    form = DailyInfoForm()
    plants = Plant.objects.all()

    processes = Process_Type.objects.all().order_by('show_order')

    if not sPlant or not sMach:
        return redirect(reverse('collection_index'))

    if sPlant:
        machs = Machine.objects.filter(plant=sPlant)

    plant = Plant.objects.get(plant_code=sPlant)
    mach = Machine.objects.get(mach_code=sMach)

    info = Daily_Prod_Info.objects.filter(plant=sPlant, mach=sMach, data_date=sData_date).first()

    if info:
        form = DailyInfoForm(instance=info, initial={'remark': info.remark.split(',')})

    if request.method == 'POST':
        if not info:  # 新增
            form = DailyInfoForm(request.POST)

            choices = get_production_choices(sData_date)
            form.fields['prod_name_a1'].choices = choices
            form.fields['prod_name_a2'].choices = choices
            form.fields['prod_name_b1'].choices = choices
            form.fields['prod_name_b2'].choices = choices

            if form.is_valid():
                tmp_form = form.save(commit=False)
                tmp_form.plant = plant
                tmp_form.mach = mach
                tmp_form.data_date = sData_date
                tmp_form.create_by = request.user
                tmp_form.update_by = request.user
                selected_options = form.cleaned_data['remark']
                tmp_form.remark = ','.join(selected_options)
                tmp_form.save()

                # 重新取得最新資料
                info = Daily_Prod_Info.objects.filter(plant=sPlant, mach=sMach, data_date=sData_date).first()
        else:  # 更新
            form = DailyInfoForm(request.POST, instance=info)

            choices = get_production_choices(sData_date)
            form.fields['prod_name_a1'].choices = choices
            form.fields['prod_name_a2'].choices = choices
            form.fields['prod_name_b1'].choices = choices
            form.fields['prod_name_b2'].choices = choices

            if form.is_valid():
                tmp_form = form.save(commit=False)
                tmp_form.update_by = request.user
                selected_options = form.cleaned_data['remark']
                tmp_form.remark = ','.join(selected_options)
                tmp_form.save()

        # =====================Start================================
        a1_product = tmp_form.prod_name_a1
        a1_size = tmp_form.prod_size_a1
        if a1_product and a1_size:
            if not Daily_Prod_Info_Head.objects.filter(data_date=sData_date, line="A1", product=a1_product,
                                                       size=a1_size, mach=mach, plant=plant):
                Daily_Prod_Info_Head.objects.create(data_date=sData_date, line="A1", product=a1_product,
                                                    size=a1_size, create_by=request.user, mach=mach, plant=plant)

        a2_product = tmp_form.prod_name_a2
        a2_size = tmp_form.prod_size_a2
        if a2_product and a2_size:
            if not Daily_Prod_Info_Head.objects.filter(data_date=sData_date, line="A2", product=a2_product,
                                                       size=a2_size, mach=mach, plant=plant):
                Daily_Prod_Info_Head.objects.create(data_date=sData_date, line="A2", product=a2_product,
                                                    size=a2_size, create_by=request.user, mach=mach, plant=plant)

        b1_product = tmp_form.prod_name_b1
        b1_size = tmp_form.prod_size_b1
        if b1_product and b1_size:
            if not Daily_Prod_Info_Head.objects.filter(data_date=sData_date, line="B1", product=b1_product,
                                                       size=b1_size, mach=mach, plant=plant):
                Daily_Prod_Info_Head.objects.create(data_date=sData_date, line="B1", product=b1_product,
                                                    size=b1_size, create_by=request.user, mach=mach, plant=plant)

        b2_product = tmp_form.prod_name_b2
        b2_size = tmp_form.prod_size_b2
        if b2_product and b2_size:
            if not Daily_Prod_Info_Head.objects.filter(data_date=sData_date, line="B2", product=b2_product,
                                                       size=b2_size, mach=mach, plant=plant):
                Daily_Prod_Info_Head.objects.create(data_date=sData_date, line="B2", product=b2_product,
                                                    size=b2_size, create_by=request.user, mach=mach, plant=plant)
        # =====================End================================

        msg = _("Update Done")

    choices = get_production_choices(sData_date)
    form.fields['prod_name_a1'].choices = choices
    form.fields['prod_name_a2'].choices = choices
    form.fields['prod_name_b1'].choices = choices
    form.fields['prod_name_b2'].choices = choices

    daily_prod_info_heads = Daily_Prod_Info_Head.objects.filter(data_date=sData_date, mach=mach, plant=plant).order_by(
        'line')

    return render(request, 'collection/daily_info_create.html', locals())

@login_required
def daily_info_head_delete(request, pk):
    if request.method == 'GET':
        head = Daily_Prod_Info_Head.objects.get(pk=pk)
        head.delete()
    return redirect(reverse('daily_info_create'))



def raw_data_api(request, data_date_start, data_date_end, process_type):
    records = []
    data_date_start = datetime.strptime(data_date_start, '%Y%m%d')
    data_date_end = datetime.strptime(data_date_end, '%Y%m%d') + timedelta(days=1)
    date_list = [data_date_start + timedelta(days=x) for x in range(0, (data_date_end - data_date_start).days)]
    process_type = process_type
    TIMES = ["00", "06", "12", "18"]
    control = ParameterDefine.objects.filter(plant="GDNBR", mach="01", process_type="ACID",
                                             parameter_name__icontains="TEMPERATURE").first()
    defines = ParameterDefine.objects.filter(process_type="ACID", parameter_name__icontains="TEMPERATURE")

    for data_date in date_list:
        for time in TIMES:
            record = {}
            record["DATA_TIME"] = datetime.strftime(data_date, '%Y/%m/%d') + " " + time + ":00"
            record["PROCESS_TYPE"] = process_type
            for define in defines:
                record["PLANT"] = define.plant.plant_code
                data = ParameterValue.objects.filter(data_date=data_date, process_type=process_type, plant=define.plant,
                                                     data_time=time, parameter_name=define.parameter_name,
                                                     mach=define.mach).first()
                record[define.mach.mach_code + "_" + define.parameter_name] = data.parameter_value if data else 0
            record["RANGE_HIGH"] = control.control_range_high
            record["BASE"] = control.base_line
            record["RANGE_LOW"] = control.control_range_low
            records.append(record)

    return JsonResponse(records, safe=False)


def get_mach_api(request):
    if request.method == 'POST':
        plant = request.POST.get('plant')
        machs = Machine.objects.filter(plant=plant)
        html = """<option value="" selected>---------</option>"""

        for mach in machs:
            html += """<option value="{value}">{name}</option>""".format(value=mach.mach_code, name=mach.mach_name)
    return JsonResponse(html, safe=False)


def test(request):
    return render(request, 'collection/test.html', locals())


@login_required
def rd_select(request):
    if request.method == 'POST':
        request.session['plant'] = request.POST.get('plant', '')
        request.session['mach'] = request.POST.get('mach', '')
        request.session['data_date'] = request.POST.get('data_date', '')
        request.session['to_date'] = request.POST.get('to_date', '')
        request.session['enable_mode'] = 'on' if request.POST.get('enable_mode') == 'on' else 'off'
        request.session['limit_mode'] = '1' if request.POST.get('limit_mode') == '1' else '0'
    else:
        request.session['enable_mode'] = request.session.get('enable_mode', 'off')
        request.session['limit_mode'] = request.session.get('limit_mode', '0')

    return (request.session.get('plant', ''), request.session.get('mach', ''),
            request.session.get('data_date', ''), request.session.get('to_date', ''),
            request.session.get('enable_mode', 'off'), request.session.get('limit_mode', '0'),
            get_language())


@login_required
def rd_report(request):
    today = datetime.today()
    sPlant, sMach, sData_date, sTo_date, sEnable_mode, sLimit_mode, lang = rd_select(request)
    if sData_date == "":
        sData_date = datetime.now().strftime('%Y-%m-%d')
    modal_message = 'Loading... Please wait while the file is being generated.'
    if sPlant != '' and sMach != '' and sData_date != '' and sTo_date != '':
        start_date = datetime.strptime(sData_date, "%Y-%m-%d")
        end_date = datetime.strptime(sTo_date, "%Y-%m-%d")
        if start_date > end_date:
            modal_message = 'Start date is after End date'
        elif (end_date - start_date).days >= 31:
            modal_message = 'Day range is bigger than 31 days'
    elif sPlant == '' or sMach == '':
        modal_message = 'Some selection empty !'

    process_type = Process_Type.objects.filter().first()
    plants = Plant.objects.all()
    machs = Machine.objects.filter(plant=sPlant) if sPlant else None
    db = vnedc_database()

    if 'GD' not in sPlant and 'LK' not in sPlant:
        sPlant = 'GDNBR'
        sMach = 'GDNBR01'
        machs = Machine.objects.filter(plant=sPlant)
    if 'GD' in sPlant:
        names = ["Acid tank 1", "Acid tank 2", "Alkaline tank 1", "Alkaline tank 2", "Latex SIDE A-1", "Latex SIDE A-2",
                 "Latex SIDE B-1", "Latex SIDE B-2", "Coagulant A", "Coagulant B", "Chlorination"]
        merge_sizes = [1, 1, 1, 1, 2, 2, 2, 2, 3, 3, 1]
        parameters = ["%", "%", "%", "%", "CN (%)", "CPF (%)", "pH Value",
                      "CN (%)", "CPF (%)", "pH Value", "TSC %", "pH Value",
                      "TSC %", "pH Value", "TSC %", "pH Value", "TSC %", "pH Value", "ppm"]
        table_data = [['Acid tank 1', ['%']], ['Acid tank 2', ['%']], ['Alkaline tank 1', ['%']],
                      ['Alkaline tank 2', ['%']],
                      ['Latex SIDE A-1', ['pH Value', 'TSC %']], ['Latex SIDE A-2', ['pH Value', 'TSC %']],
                      ['Latex SIDE B-1', ['pH Value', 'TSC %']], ['Latex SIDE B-2', ['pH Value', 'TSC %']],
                      ['Coagulant A', ['CN (%)', 'CPF (%)', 'pH Value']],
                      ['Coagulant B', ['CN (%)', 'CPF (%)', 'pH Value']],
                      ['Chlorination', ['ppm']], ['Độ ẩm', ['%']], ['Hàm lượng bột', ['mg/gloves']], ['PreLeach', ['TDS']]]


    elif 'LK' in sPlant:
        table_data = [['Acid tank 1', ['%']], ['Acid tank 2', ['%']], ['Alkaline tank 1', ['%']],
                      ['Alkaline tank 2', ['%']],
                      ['Latex 1', ['pH Value', 'TSC %']], ['Latex 2', ['pH Value', 'TSC %']],
                      ['Coagulant A', ['CN (%)', 'CPF (%)', 'pH Value']],
                      ['Coagulant B', ['CN (%)', 'CPF (%)', 'pH Value']],
                      ['Chlorination', ['ppm']], ['Độ ẩm', ['%']], ['Hàm lượng bột', ['mg/gloves']], ['PreLeach', ['TDS']]]

    try:
        sql_mach = f"""
                    SELECT data_date, mach_id
                    FROM [VNEDC].[dbo].[collection_daily_prod_info_head] where data_date = '{sData_date}' 
                    order by mach_id"""

        machines = db.select_sql_dict(sql_mach)
        machines_list = [machine['mach_id'] for machine in machines]
        machines_list = list(set(machines_list))
        machines_no_list = sorted([machine[-2:] for machine in machines_list])
        data_time = ['00', '06', '12', '18']
        sql_No = f"""
                    SELECT product
                    FROM [VNEDC].[dbo].[collection_daily_prod_info_head]
                    where data_date = '{sData_date}' and mach_id = '{sMach}'
                    """
        result02 = db.select_sql_dict(sql_No)
        result = '/'.join(item['product'] for item in result02)
        itemNo_display = ' ,'.join(sorted(list(set([result['product'] for result in result02]))))
        itemNo = re.findall(r'\d+', result)[-1] if len(result) > 0 else 0

        sql = f"""
                    SELECT pd.process_type_id, pd.parameter_name, pc.control_range_low, pc.control_range_high,
                    MAX(CASE WHEN pv.data_time = '00' THEN pv.parameter_value ELSE NULL END) AS at_00,
                    MAX(CASE WHEN pv.data_time = '06' THEN pv.parameter_value ELSE NULL END) AS at_06,
                    MAX(CASE WHEN pv.data_time = '12' THEN pv.parameter_value ELSE NULL END) AS at_12,
                    MAX(CASE WHEN pv.data_time = '18' THEN pv.parameter_value ELSE NULL END) AS at_18,
                    pd.mach_id, 
                    Case 
                        When CHARINDEX('00', pd.input_time) > 0 THEN 1
                        else 0
                    end as time_00,
                    Case 
                        When CHARINDEX('06', pd.input_time) > 0 THEN 1
                        else 0
                    end as time_06,
                    Case 
                        When CHARINDEX('12', pd.input_time) > 0 THEN 1
                        else 0
                    end as time_12,
                    Case 
                        When CHARINDEX('18', pd.input_time) > 0 THEN 1
                        else 0
                    end as time_18
                    FROM [VNEDC].[dbo].[collection_parameterdefine] pd
                    left join [VNEDC].[dbo].[collection_parametervalue]  pv
                    on pv.plant_id = pd.plant_id and pv.mach_id = pd.mach_id and pd.process_type_id = pv.process_type and pd.parameter_name = pv.parameter_name and pv.data_date = '{sData_date}'
                    left join [VNEDC].[dbo].[collection_lab_parameter_control] pc
                    on item_no = {itemNo} and pc.mach_id = pd.mach_id and pc.process_type = pd.process_type_id  and CHARINDEX(pc.parameter_name, pd.parameter_name) > 0
                    where pd.plant_id = '{sPlant}' and pd.mach_id = '{sMach}'  and
                    ((pd.process_type_id = 'COAGULANT' AND (pd.parameter_name LIKE '%CPF%' OR pd.parameter_name LIKE '%pH%' OR pd.parameter_name LIKE '%CONCENTRATION%'))
                    OR (pd.process_type_id = 'LATEX' AND (pd.parameter_name LIKE '%TSC%' OR pd.parameter_name LIKE '%pH%'))
                    OR (pd.process_type_id = 'ACID' AND pd.parameter_name LIKE '%CONCENTRATION%')
                    OR (pd.process_type_id = 'ALKALINE' AND pd.parameter_name LIKE '%CONCENTRATION%')
                    OR (pd.process_type_id = 'CHLORINE' AND pd.parameter_name = 'CONCENTRATION')
                    OR (pd.process_type_id = 'OTHER' AND pd.parameter_name = 'MOISTURE_CONTENT')
                    OR (pd.process_type_id = 'OTHER' AND pd.parameter_name = 'POWDER_CONTENT')
                    OR (pd.process_type_id = 'PRELEACH' AND pd.parameter_name = 'T5_TDS'))
                    GROUP BY 
                        pd.process_type_id, 
                        pd.parameter_name,
                        pd.mach_id,
                        pc.control_range_low, 
                        pc.control_range_high,
                        pd.input_time
                    ORDER BY 
                        CASE
                            WHEN pd.process_type_id = 'ACID' THEN 1
                            WHEN pd.process_type_id = 'ALKALINE' THEN 2
                            WHEN pd.process_type_id = 'LATEX' THEN 3
                            WHEN pd.process_type_id = 'COAGULANT' THEN 4
                            WHEN pd.process_type_id = 'CHLORINE' THEN 5
                            WHEN pd.process_type_id = 'OTHER' THEN 6
                            WHEN pd.process_type_id = 'PRELEACH' THEN 7
                        ELSE 8 END,
                        pd.process_type_id,   
                        pd.parameter_name ASC;
                """
        results = db.select_sql_dict(sql)
        limit = []
        data = []
        mode0 = 0
        mode6 = 0
        mode12 = 0
        mode18 = 0
        for result in results:
            limit_range = f"{result['control_range_low']} ~ {result['control_range_high']}" if str(
                result['control_range_high']) != 'None' else ' '
            limit_low = float(result['control_range_low']) if result['control_range_low'] is not None else 0
            limit_high = float(result['control_range_high']) if result['control_range_low'] is not None else 10000
            at_0 = float(result['at_00']) if result['at_00'] is not None else -1
            at_6 = float(result['at_06']) if result['at_06'] is not None else -1
            at_12 = float(result['at_12']) if result['at_12'] is not None else -1
            at_18 = float(result['at_18']) if result['at_18'] is not None else -1
            time_0 = int(result['time_00'])
            time_6 = int(result['time_06'])
            time_12 = int(result['time_12'])
            time_18 = int(result['time_18'])
            is_date = 1 if datetime.strptime(sData_date, '%Y-%m-%d') > today else 0
            limit.append(limit_range)
            data.append([limit_low, limit_high, at_0, at_6, at_12, at_18, time_0, time_6, time_12, time_18])
    except:
        pass
    return render(request, 'collection/rd_report.html', locals())

def rd_message(request):
    if request.method == 'POST':
        sPlant = request.POST.get('plant')
        sData_date = request.POST.get('data_date')
        select1_value = request.POST.get('select1')
        select2_value = request.POST.get('select2')
        at_time_value = request.POST.get('at_time')
        action = request.POST.get('action')
        image = request.POST.get('image')
        confirm = rd_report_confirm(select1_value, select2_value, at_time_value, sPlant, sData_date)
        # message = rd_report_message(select1_value, select2_value, at_time_value, sPlant, sData_date)
        if action == "send_wecom":
            base64_str = str(image).split(',')[1]
            image_data = base64.b64decode(base64_str)
            md5 = hashlib.md5(image_data).hexdigest()
            image_base64 = base64.b64encode(image_data).decode('utf-8')
            path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            print(path)
            wecom_file = os.path.join(path, "static", "wecom", "key_param.config")
            key = ''
            if os.path.exists(wecom_file):
                with open(wecom_file, 'r') as file:
                    key = file.read().strip()

            send_message(key, image_base64, md5)
        result = {"result": confirm}

        return JsonResponse(result, safe=False)

def send_message(key, image_base64, md5_hash):
    try:
        headers = {'Content-Type': 'application/json; charset=utf-8'}
        data = {
            "msgtype": "image",
            "image": {
                "base64": image_base64,
                "md5": md5_hash
            }
        }
        headers = {'Content-Type': 'application/json'}
        response = requests.post(key, json=data, headers=headers)
        return response.json()
    except:
        pass

def rd_report_confirm(select1, select2, at_time, plant, date):
    try:
        select_mach1 = f"{plant}{select1}"
        select_mach2 = f"{plant}{select2}"

        db = vnedc_database()
        sql01 = f"""SELECT mach_id
                           FROM [VNEDC].[dbo].[collection_daily_prod_info_head] where data_date = '{date}' 
                           and mach_id between '{select_mach1}' and '{select_mach2}'
                           order by mach_id"""
        machines_result = db.select_sql_dict(sql01)
        machines_list = [machine['mach_id'] for machine in machines_result]
        machines = sorted(list(set(machines_list)))
        if len(machines) > 1:
            machines_text = "'" + "','".join(
                machines) + "'"  # 'GDNBR01','GDNBR02','GDNBR03','GDNBR04','GDNBR05','GDNBR06','GDNBR07'
        else:
            machines_text = f"'{machines[0]}'"
        text = f'SELECT pd.process_type_id, pd.parameter_name'
        item_no = []
        for machine in machines:
            sql02 = f"""SELECT product
                            FROM [VNEDC].[dbo].[collection_daily_prod_info_head]
                            where data_date = '{date}' and mach_id = '{machine}'
                        """
            itemNo_result = db.select_sql_dict(sql02)
            result = '/'.join(item['product'] for item in itemNo_result)
            itemNo = int(re.findall(r'\d+', result)[0] if len(result) > 0 else 0)
            item_no.append(itemNo)
            text += f""",MAX(CASE WHEN pv.data_time = '{at_time}' AND pv.mach_id = '{machine}' THEN pv.parameter_value ELSE NULL END) AS {machine},
                             CASE 
                                 WHEN MAX(CASE WHEN pc.item_no = {itemNo} and pc.mach_id = '{machine}' THEN pc.control_range_low ELSE NULL END) > MAX(CASE WHEN pv.data_time = '{at_time}' AND pv.mach_id = '{machine}' THEN pv.parameter_value ELSE NULL END) THEN -1
                                 WHEN MAX(CASE WHEN pc.item_no = {itemNo} and pc.mach_id = '{machine}' THEN pc.control_range_high ELSE NULL END) < MAX(CASE WHEN pv.data_time = '{at_time}' AND pv.mach_id = '{machine}' THEN pv.parameter_value ELSE NULL END) THEN 1	
                                 ELSE 0
                             END as {machine}_mode,
                             CASE 
                                 WHEN MAX(CASE 
                                     WHEN pd.mach_id = '{machine}' AND CHARINDEX('{at_time}', pd.input_time) > 0 THEN 1 
                                     ELSE 0 
                                     END) = 1 THEN 1 
                                 ELSE 0 
                             END AS {machine}_bg
                                """
        text += f"""FROM [VNEDC].[dbo].[collection_parameterdefine] pd
                        LEFT JOIN [VNEDC].[dbo].[collection_parametervalue] pv
                        ON pv.plant_id = pd.plant_id
                        AND pv.mach_id IN ({machines_text})
                        AND pd.process_type_id = pv.process_type 
                        AND pd.parameter_name = pv.parameter_name 
                        AND pv.data_date = '{date}'
                        LEFT JOIN [VNEDC].[dbo].[collection_lab_parameter_control] pc"""
        item_no_text = list(set(item_no))
        if len(item_no_text) > 1:
            item_no_text = tuple(
                int(value) for value in item_no_text)  # Convert to tuple of integers for multiple items
            text += f" ON pc.item_no IN {item_no_text}"
        else:
            item_no_text = (item_no_text[0])
            text += f" ON pc.item_no = {item_no_text}"
        text += f"""AND pc.mach_id = pv.mach_id
                        AND pc.process_type = pd.process_type_id
                        AND CHARINDEX(pc.parameter_name, pd.parameter_name) > 0
                        WHERE pd.plant_id = '{plant}'  AND
                                (
                                (pd.process_type_id = 'COAGULANT' AND (pd.parameter_name LIKE '%CPF%' OR pd.parameter_name LIKE '%pH%' OR pd.parameter_name LIKE '%CONCENTRATION%'))
                                OR (pd.process_type_id = 'LATEX' AND (pd.parameter_name LIKE '%TSC%' OR pd.parameter_name LIKE '%pH%'))
                                OR (pd.process_type_id = 'ACID' AND pd.parameter_name LIKE '%CONCENTRATION%')
                                OR (pd.process_type_id = 'ALKALINE' AND pd.parameter_name LIKE '%CONCENTRATION%')
                                OR (pd.process_type_id = 'CHLORINE' AND pd.parameter_name = 'CONCENTRATION')
                                OR (pd.process_type_id = 'OTHER' AND pd.parameter_name = 'MOISTURE_CONTENT')
                                OR (pd.process_type_id = 'OTHER' AND pd.parameter_name = 'POWDER_CONTENT')
                                )
                        GROUP BY 
                            pd.process_type_id, 
                            pd.parameter_name,
                            pd.input_time
                        ORDER BY
                            CASE
                                WHEN pd.process_type_id = 'ACID' THEN 1
                                WHEN pd.process_type_id = 'ALKALINE' THEN 2
                                WHEN pd.process_type_id = 'LATEX' THEN 3
                                WHEN pd.process_type_id = 'COAGULANT' THEN 4
                                WHEN pd.process_type_id = 'CHLORINE' THEN 5
                            ELSE 6 END,
                            pd.process_type_id,   
                            pd.parameter_name ASC;"""

        data_table = db.select_sql_dict(text)
        for row in data_table:
            # if row['process_type_id'] == 'ACID' or row['process_type_id'] == 'ALKALINE' or row[
            #     'process_type_id'] == 'CHLORINE':
            #     row['parameter_name'] = ''
            # if row['process_type_id'] == 'LATEX':
            #     row['parameter_name'] = f" {row['parameter_name'][0]}{row['parameter_name'][3:]}"
            # if row['process_type_id'] == 'OTHER':
            #     row['process_type_id'] = ''
            #     row['parameter_name'] = (str(row['parameter_name']).split('_'))[0]
            # if row['process_type_id'] == 'COAGULANT':
            #     if 'CONCENTRATION' in row['parameter_name']:
            #         row['parameter_name'] = f" {row['parameter_name'][:2]}CN"
            #     else:
            #         row['parameter_name'] = f" {row['parameter_name']}"
            row['parameter_name'] = str(row['parameter_name']).replace('PH', 'pH')

        first_row = [[f"Ngày: {date} {at_time}:00"]]
        for value in machines:
            first_row.append([value[2:], 99, 1])
        data_rows = []
        data_rows.append(first_row)
        for rows in data_table:
            # item0 = f"{rows['process_type_id']} {str(rows['parameter_name'])[:str(rows['parameter_name']).rfind('_')]}"
            item0 = f" {rows['process_type_id']}_{str(rows['parameter_name'])}"
            values = [value for key, value in rows.items() if 'GDNBR' in key]
            grouped_values = [item0] + [values[i:i + 3] for i in range(0, len(values), 3)]
            data_rows.append(grouped_values)
        return data_rows
    except:
        pass


def rd_report_message(select1, select2, at_time, plant, date):
    try:
        db = vnedc_database()
        select_mach1 = f"{plant}{select1}"
        select_mach2 = f"{plant}{select2}"
        message = ''
        sql01 = f"""SELECT mach_id
                        FROM [VNEDC].[dbo].[collection_daily_prod_info_head] where data_date = '{date}' 
                        and mach_id between '{select_mach1}' and '{select_mach2}'
                        order by mach_id"""
        machines_result = db.select_sql_dict(sql01)
        machines_list = [machine['mach_id'] for machine in machines_result]
        machines = sorted(list(set(machines_list)))
        for machine in machines:
            sql02 = f"""SELECT product
                            FROM [VNEDC].[dbo].[collection_daily_prod_info_head]
                            where data_date = '{date}' and mach_id = '{machine}'
                        """
            itemNo_result = db.select_sql_dict(sql02)
            result = '/'.join(item['product'] for item in itemNo_result)
            itemNo = re.findall(r'\d+', result)[0] if len(result) > 0 else 0

            sql03 = f"""SELECT pv.mach_id, pv.process_type, pv.parameter_name, pv.parameter_value, pc.control_range_low ,pc.control_range_high,
                            case 
                            when pv.parameter_value < pc.control_range_low then -1
                            when pv.parameter_value > pc.control_range_high then 1
                            else 0
                            end as mode
                            FROM [VNEDC].[dbo].[collection_parametervalue] pv
                            left join [VNEDC].[dbo].[collection_lab_parameter_control] pc
                            on  pc.item_no = {itemNo} and pc.process_type = pv.process_type and pc.mach_id = pv.mach_id and CHARINDEX(pc.parameter_name, pv.parameter_name) > 0 
                            where pv.plant_id = '{plant}' and pv.mach_id = '{machine}' and  pv.data_date = '{date}' and pv.data_time = '{at_time}' 
                            and ((pv.process_type = 'ACID' and pv.parameter_name LIKE '%CONCENTRATION%') or (pv.process_type = 'ALKALINE' and pv.parameter_name LIKE '%CONCENTRATION%')
                            or (pv.process_type = 'COAGULANT' and (pv.parameter_name like '%CPF%' or pv.parameter_name like '%pH%' or pv.parameter_name like '%CONCENTRATION%'))
                            or (pv.process_type = 'LATEX' and (pv.parameter_name like '%TSC%' or pv.parameter_name like '%pH%')))
                            order by process_type, parameter_name
                            """
            result03 = db.select_sql_dict(sql03)
            acid_list = [[], []]
            alkaline_list = [[], []]
            latex_ph = [[], []]
            latex_tsc = [[], []]
            coagulant_a = [[], []]
            coagulant_b = [[], []]
            text_01, text_02, text_03, text_04, text_05, text_06, text_07 = '', '', '', '', '', '', ''
            for result in result03:
                if result['process_type'] == 'ACID':
                    if float(result['mode']) == -1:
                        acid_list[0].append(str(result['parameter_name'])[1])
                    elif float(result['mode']) == 1:
                        acid_list[1].append(str(result['parameter_name'])[1])

                if len(acid_list[0]) > 0:
                    text_01 = '+ Acid: ' + ', '.join(acid_list[0]) + ' thấp'
                if len(acid_list[1]) > 0:
                    if len(acid_list[0]) > 0:
                        text_01 += '- ' + ', '.join(acid_list[1]) + ' cao'
                    else:
                        text_01 = '+ Acid: ' + ', '.join(acid_list[1]) + ' cao'

                if result['process_type'] == 'ALKALINE':
                    if float(result['mode']) == -1:
                        alkaline_list[0].append(str(result['parameter_name'])[1])
                    elif float(result['mode']) == 1:
                        alkaline_list[1].append(str(result['parameter_name'])[1])

                if len(alkaline_list[0]) > 0:
                    text_02 = '+ Alkaline: ' + ', '.join(alkaline_list[0]) + ' thấp'
                if len(alkaline_list[1]) > 0:
                    if len(alkaline_list[0]) > 0:
                        text_02 += ' - ' + ', '.join(alkaline_list[1]) + ' cao'
                    else:
                        text_02 = '+ Alkaline: ' + ', '.join(alkaline_list[1]) + ' cao'

                if result['process_type'] == 'LATEX':
                    if 'PH' in result['parameter_name']:
                        if result['mode'] == -1:
                            latex_ph[0].append(f"{result['parameter_name'][0]}{result['parameter_name'][3]}")
                        elif result['mode'] == 1:
                            latex_ph[1].append(f"{result['parameter_name'][0]}{result['parameter_name'][3]}")

                    if len(latex_ph[0]) > 0:
                        text_03 = '+ Latex PH: ' + ', '.join(latex_ph[0]) + ' thấp'
                    if len(latex_ph[1]) > 0:
                        if len(latex_ph[0]) > 0:
                            text_03 += ' - ' + ', '.join(latex_ph[1]) + ' cao'
                        else:
                            text_03 = '+ Latex PH: ' + ', '.join(latex_ph[1]) + ' cao'

                    if 'TSC' in result['parameter_name']:
                        if result['mode'] == -1:
                            latex_tsc[0].append(f"{result['parameter_name'][0]}{result['parameter_name'][3]}")
                        elif result['mode'] == 1:
                            latex_tsc[1].append(f"{result['parameter_name'][0]}{result['parameter_name'][3]}")

                    if len(latex_tsc[0]) > 0:
                        text_04 = '+ Latex TSC: ' + ', '.join(latex_tsc[0]) + ' thấp'
                    if len(latex_tsc[1]) > 0:
                        if len(latex_tsc[0]) > 0:
                            text_04 += ' - ' + ', '.join(latex_tsc[1]) + ' cao'
                        else:
                            text_04 = '+ Latex TSC: ' + ', '.join(latex_tsc[1]) + ' cao'

                if result['process_type'] == 'COAGULANT':
                    if result['parameter_name'][0] == 'A':
                        if float(result['mode']) == -1:
                            coagulant_a[0].append(
                                'CN' if result['parameter_name'][2:] == 'CONCENTRATION' else result['parameter_name'][
                                                                                             2:])
                        elif float(result['mode']) == 1:
                            coagulant_a[1].append(
                                'CN' if result['parameter_name'][2:] == 'CONCENTRATION' else result['parameter_name'][
                                                                                             2:])

                    if len(coagulant_a[0]) > 0:
                        text_05 = '+ Coagulant A: ' + ', '.join(coagulant_a[0]) + ' thấp'
                    if len(coagulant_a[1]) > 0:
                        if len(coagulant_a[0]) > 0:
                            text_05 += ' - ' + ', '.join(coagulant_a[1]) + ' cao'
                        else:
                            text_05 = '+ Coagulant A: ' + ', '.join(coagulant_a[1]) + ' cao'

                    if result['parameter_name'][0] == 'B':
                        if result['mode'] == -1:
                            coagulant_b[0].append(
                                'CN' if result['parameter_name'][2:] == 'CONCENTRATION' else result['parameter_name'][
                                                                                             2:])
                        elif result['mode'] == 1:
                            coagulant_b[1].append(
                                'CN' if result['parameter_name'][2:] == 'CONCENTRATION' else result['parameter_name'][
                                                                                             2:])

                    if len(coagulant_b[0]) > 0:
                        text_06 = '+ Coagulant B: ' + ', '.join(coagulant_b[0]) + ' thấp'
                    if len(coagulant_b[1]) > 0:
                        if len(coagulant_b[0]) > 0:
                            text_06 += ' - ' + ', '.join(coagulant_b[1]) + ' cao'
                        else:
                            text_06 = '+ Coagulant B: ' + ', '.join(coagulant_b[1]) + ' cao'
            text_list = [machine, text_01, text_02, text_03, text_04, text_05, text_06, text_07]
            text = '\n'.join([t for t in text_list if t])
            if len(text) > 8:
                message += text + '\n'
        return message
    except:
        pass

def generate_excel_file_big(request):
    sPlant, sMach, sData_date, sTo_date, sEnable_mode, sLimit_mode, lang = rd_select(request)
    if sEnable_mode == 'off':
        sTo_date == ''
    workbook = Workbook()
    worksheet = workbook.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    download_code = 0
    if sPlant != '' and sMach != '' and sData_date != '':
        if sTo_date != '':
            start_date = datetime.strptime(sData_date, "%Y-%m-%d")
            end_date = datetime.strptime(sTo_date, "%Y-%m-%d")
            if start_date == end_date:
                filename = f"DATA-{sPlant}-{sMach}-{sData_date}.xlsx"
                plant_rp, mach_rp, date_rp, to_rp = sPlant, sMach, sData_date, sTo_date
                download_code = 1
            elif start_date < end_date and (end_date - start_date).days < 31:
                filename = f"DATA-{sPlant}-{sMach}-{sData_date}-to-{sTo_date}.xlsx"
                plant_rp, mach_rp, date_rp, to_rp = sPlant, sMach, sData_date, sTo_date
                download_code = 1
        elif sTo_date == '':
            start_date = datetime.strptime(sData_date, "%Y-%m-%d")
            end_date = start_date
            filename = f"DATA-{sPlant}-{sMach}-{sData_date}.xlsx"
            plant_rp, mach_rp, date_rp, to_rp = sPlant, sMach, sData_date, sTo_date
            download_code = 1

    max_row = 30
    if 'GDNBR' in sPlant:
        names = ["Acid tank 1", "Acid tank 2", "Alkaline tank 1", "Alkaline tank 2", "Coagulant A",
                 "Coagulant B", "Latex SIDE A-1", "Latex SIDE A-2", "Latex SIDE B-1", "Latex SIDE B-2", "Chlorination", 'Độ ẩm', 'Hàm lượng bột', 'PreLeach']
        merge_sizes = [1, 1, 1, 1, 3, 3, 2, 2, 2, 2, 1, 1, 1, 1]
        parameters = ["%", "%", "%", "%", "CN (%)", "CPF (%)", "pH Value",
                      "CN (%)", "CPF (%)", "pH Value", "TSC %", "pH Value",
                      "TSC %", "pH Value", "TSC %", "pH Value", "TSC %", "pH Value", "ppm", "%", "mg/gloves", "TDS"]
    elif 'LK' in sPlant:
        names = ["Acid tank 1", "Acid tank 2", "Alkaline tank 1", "Alkaline tank 2", "Coagulant A", "Coagulant B",
                 "Latex 1", "Latex 2", "Chlorination", 'Độ ẩm', 'Hàm lượng bột', "PreLeach"]
        merge_sizes = [1, 1, 1, 1, 3, 3, 2, 2, 1, 1, 1, 1]
        parameters = ["%", "%", "%", "%", "CN (%)", "CPF (%)", "pH Value",
                      "CN (%)", "CPF (%)", "pH Value", "TSC %", "pH Value",
                      "TSC %", "pH Value", "ppm", "%", "mg/gloves", "TDS"]
        max_row = 26

    if download_code == 1:
        for row in range(5, max_row + 1):
            worksheet.row_dimensions[row].height = 25
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 20
        # worksheet.column_dimensions['C'].width = 20
        worksheet.row_dimensions[1].height = 45
        worksheet.row_dimensions[2].height = 30
        worksheet.row_dimensions[3].height = 30
        worksheet.row_dimensions[4].height = 30

        def set_merged_cell(range_str, value, alignment="center", fill_color="F1FF18"):
            worksheet.merge_cells(range_str)
            cell = worksheet[range_str.split(':')[0]]
            cell.value = value
            cell.alignment = Alignment(horizontal=alignment, vertical="center")
            cell.font = Font(bold=True, color="333333")
            cell.fill = PatternFill("solid", fgColor=fill_color)

        def set_merged_cell2(range_str, value, alignment="center", fill_color="FDE9D9"):
            worksheet.merge_cells(range_str)
            cell = worksheet[range_str.split(':')[0]]
            cell.value = value
            cell.alignment = Alignment(horizontal=alignment, vertical="center")
            cell.font = Font(bold=True, color="333333")
            cell.fill = PatternFill("solid", fgColor=fill_color)

        set_merged_cell('A6:A8', "ITEM")
        set_merged_cell('B6:B8', "Parameters")
        set_merged_cell('C6:C8', "Specs.")
        set_merged_cell('D6:G6', "Sample Time")

        start_row = 9
        for name, size in zip(names, merge_sizes):
            end_row = start_row + size - 1
            worksheet.merge_cells(f"A{start_row}:A{end_row}")
            cell = worksheet[f"A{start_row}"]
            cell.value = name
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=3)
            start_row = end_row + 1
        for row_num, param in enumerate(parameters, start=9):
            worksheet[f'B{row_num}'] = param
            worksheet[f'B{row_num}'].alignment = Alignment(horizontal="center", vertical="center")

        days_diff = (end_date - start_date).days + 1
        col_end = 7

        product = []
        for day_offset in range(days_diff):
            current_date = start_date + timedelta(days=day_offset)
            col_start = 3 + (day_offset * 5)
            col_end = col_start + 4
            if sPlant == 'GDNBR':
                sql = f"""
                    SELECT product
                    FROM [VNEDC].[dbo].[collection_daily_prod_info_head]
                    where data_date = '{str((current_date).strftime('%Y-%m-%d'))}' and mach_id = '{sMach}'
                """
                db = vnedc_database()
                results = db.select_sql_dict(sql)
                result = '/'.join(item['product'] for item in results)
                if len(result) > 0:
                    result = result
                else:
                    result = ''
            else:
                result = ''
            product.append(result)
            # print(result)

            worksheet.column_dimensions[f'{get_column_letter(col_start)}'].width = 17
            set_merged_cell(f'{get_column_letter(col_start)}6:{get_column_letter(col_start)}8', "Specs.")
            set_merged_cell2(f'{get_column_letter(col_start)}2:{get_column_letter(col_end)}2', f"{plant_rp}")
            set_merged_cell2(f'{get_column_letter(col_start)}3:{get_column_letter(col_end)}3', f"{mach_rp}")
            set_merged_cell2(f'{get_column_letter(col_start)}4:{get_column_letter(col_end)}4',
                             f"{current_date.strftime('%Y-%m-%d')}")
            set_merged_cell2(f'{get_column_letter(col_start)}5:{get_column_letter(col_end)}5', result)
            set_merged_cell(f'{get_column_letter(col_start+1)}6:{get_column_letter(col_end)}6', "Sample Time")
            sample_times = ["1", "2", "3", "4"]
            sample_hours = ["0h", "6h", "12h", "18h"]

            for i, col in enumerate(
                    [get_column_letter(col_start + 1), get_column_letter(col_start + 2), get_column_letter(col_end - 1),
                     get_column_letter(col_end)]):
                worksheet[f'{col}7'] = sample_times[i]
                worksheet[f'{col}8'] = sample_hours[i]
                worksheet[f'{col}7'].alignment = worksheet[f'{col}8'].alignment = Alignment(horizontal="center")
                worksheet[f'{col}7'].fill = worksheet[f'{col}8'].fill = PatternFill("solid", fgColor="F1FF18")
            if 'GD' in sPlant:
                defines = ParameterDefine.objects.filter(
                    plant=sPlant,
                    mach=sMach,
                    process_type__in=['ACID', 'ALKALINE', 'LATEX', 'COAGULANT', 'CHLORINE', 'OTHER', 'PRELEACH'],
                    parameter_name__in=['T1_CONCENTRATION', 'T2_CONCENTRATION', 'A_T1_TSC',
                                        'A_T1_PH', 'A_T2_TSC', 'A_T2_PH', 'B_T1_TSC',
                                        'B_T1_PH', 'B_T2_TSC', 'B_T2_PH', 'A_CPF',
                                        'A_PH', 'A_CONCENTRATION', 'B_CPF',
                                        'B_PH', 'B_CONCENTRATION', 'CONCENTRATION', 'MOISTURE_CONTENT', 'POWDER_CONTENT', 'T5_TDS']

                )
            elif 'LK' in sPlant:
                defines = ParameterDefine.objects.filter(
                    plant=sPlant,
                    mach=sMach,
                    process_type__in=['ACID', 'ALKALINE', 'LATEX', 'COAGULANT', 'CHLORINE', 'OTHER',  'PRELEACH'],
                    parameter_name__in=['T1_CONCENTRATION', 'T2_CONCENTRATION', 'T1_TSC',
                                        'T1_PH', 'T2_TSC', 'T2_PH', 'A_CPF',
                                        'A_PH', 'A_CONCENTRATION', 'B_CPF',
                                        'B_PH', 'B_CONCENTRATION', 'CONCENTRATION', 'MOISTURE_CONTENT', 'POWDER_CONTENT', 'T5_TDS']
                )
            start_row = 9
            defines = list(defines)
            defines.sort(key=lambda x: x.parameter_name == 'T5_TDS')
            for define in defines:
                values = ParameterValue.objects.filter(
                    plant=sPlant,
                    mach=sMach,
                    data_date=current_date.strftime('%Y-%m-%d'),
                    process_type=define.process_type.process_code,
                    parameter_name=define.parameter_name
                )

                for value in values:
                    # Determine the column based on data_time
                    if value.data_time == '00':
                        worksheet[f'{get_column_letter(col_start+1)}{start_row}'] = value.parameter_value
                    elif value.data_time == '06':
                        worksheet[f'{get_column_letter(col_start + 2)}{start_row}'] = value.parameter_value
                    elif value.data_time == '12':
                        worksheet[f'{get_column_letter(col_start + 3)}{start_row}'] = value.parameter_value
                    elif value.data_time == '18':
                        worksheet[f'{get_column_letter(col_start + 4)}{start_row}'] = value.parameter_value
                start_row += 1

        if int(sLimit_mode) == 1:
            try:
                days_diff = (end_date - start_date).days + 1
                for day_offset in range(days_diff):
                    col_start = 3 + (day_offset * 5)
                    control_limit = Lab_Parameter_Control.objects.filter(
                        plant=sPlant,
                        mach=sMach,
                        item_no=re.findall(r'\d+', product[day_offset])[0] if len(product[day_offset]) > 0 else
                        re.findall(r'\d+', product[day_offset - 1])[0])
                    low_value, high_value = [], []
                    for i in range(len(control_limit)):
                        high_value.append(control_limit[i].control_range_high)
                        low_value.append(control_limit[i].control_range_low)

                    low_limit, high_limit = [0] * 19, [0] * 19
                    for i in ([5, 6, 8, 9] + list(range(10, 18))):
                        if i in [5, 6, 8, 9]:
                            if i == 5 or i == 8:
                                low_limit[i], high_limit[i] = low_value[3], high_value[3]
                            elif i == 6 or i == 9:
                                low_limit[i], high_limit[i] = low_value[4], high_value[4]
                        else:
                            if i % 2 == 0:
                                low_limit[i], high_limit[i] = low_value[0], high_value[0]
                            else:
                                low_limit[i], high_limit[i] = low_value[1], high_value[1]

                    range_limit = []
                    for i in range(len(high_limit)):
                        range_limit.append(f'{low_limit[i]} ~ {high_limit[i]}')

                    start_row = 9

                    for i in range(len(high_limit)):
                        if high_limit[i] != 0:
                            worksheet[f'{get_column_letter(col_start)}{start_row}'] = range_limit[i]
                            worksheet[f'{get_column_letter(col_start)}{start_row}'].alignment = Alignment(
                                vertical="center", horizontal="right")
                        start_row += 1

                    for i in range(len(high_limit)):
                        start_row = i + 9
                        if high_limit[i] != 0:
                            for day_offset in range(days_diff):
                                current_start_col = 3 + (day_offset * 5)
                                for col in range(current_start_col + 1, current_start_col + 5):
                                    cell = worksheet[f'{get_column_letter(col)}{start_row}']
                                    if cell.value and float(cell.value):
                                        if float(cell.value) > float(high_limit[i]):
                                            cell.fill = PatternFill("solid", fgColor='FF0000')
                                        elif float(cell.value) < float(low_limit[i]):
                                            cell.fill = PatternFill("solid", fgColor='FF0000')
                        start_row += 1
            except Exception as e:
                print(f"An error occurred: {e}")

        def set_header_row(row, col, value, merged_cells=None, fill_color="FDE9D9"):
            if merged_cells:
                worksheet.merge_cells(merged_cells)
            worksheet.row_dimensions[row].height = 30
            cell = worksheet[f'{col}{row}']
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(size=14, bold=False)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        header_data = [('A2', 'Plant :', plant_rp), ('A3', 'Machine :', mach_rp), ('A4', 'Date :', date_rp),
                       ('A5', 'Product', 'Test')]
        for col, label, value in header_data:
            row = int(col[1:])
            set_header_row(row, col[0], label)
        worksheet.merge_cells('B2:B5')
        fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')
        for row in worksheet.iter_rows(min_row=2, max_row=5, min_col=2, max_col=3):
            for cell in row:
                cell.fill = fill
        worksheet.merge_cells(f'A1:{get_column_letter(col_end)}1')
        worksheet.row_dimensions[1].height = 45
        merged_cell = worksheet['A1']
        merged_cell.value = 'Daily Report'
        merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        merged_cell.font = Font(size=18, bold=True)
        merged_cell.fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")

        left_top_thick = Border(left=Side(style='thick'), right=Side(style='thin'),
                                top=Side(style='thick'), bottom=Side(style='thin'))
        right_top_thick = Border(left=Side(style='thin'), right=Side(style='thick'),
                                 top=Side(style='thick'), bottom=Side(style='thin'))
        left_bot_thick = Border(left=Side(style='thick'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thick'))
        right_bot_thick = Border(left=Side(style='thin'), right=Side(style='thick'),
                                 top=Side(style='thin'), bottom=Side(style='thick'))
        right_thick = Border(left=Side(style='thin'), right=Side(style='thick'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        left_thick = Border(left=Side(style='thick'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        top_thick = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thick'), bottom=Side(style='thin'))
        bot_thick = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thick'))
        middle = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for row in worksheet.iter_rows(min_row=5, max_row=max_row, min_col=4, max_col=7):
            for cell in row:
                if cell.row == 6 and cell.column == 4:
                    cell.border = left_top_thick
                elif cell.row == 6 and cell.column == 7:
                    cell.border = right_top_thick
                elif cell.row == max_row and cell.column == 4:
                    cell.border = left_bot_thick
                elif cell.row == max_row and cell.column == 7:
                    cell.border = right_bot_thick
                elif cell.column == 4:
                    cell.border = left_thick
                elif cell.column == 7:
                    cell.border = right_thick
                elif cell.row == 6:
                    cell.border = top_thick
                elif cell.row == max_row:
                    cell.border = bot_thick
                else:
                    cell.border = middle

        for row in worksheet.iter_rows(min_row=6, max_row=8, min_col=1, max_col=col_end):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(size=14)

        for row in worksheet.iter_rows(min_row=9, max_row=max_row, min_col=1, max_col=col_end):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(size=13)
        for row in worksheet.iter_rows(min_row=2, max_row=5, min_col=1, max_col=col_end):
            for cell in row:
                cell.font = Font(size=14)
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.isdigit():
                    cell.value = int(cell.value)

        def thick_boder_table(ws, min_row, max_row, min_col, max_col):
            thick = Side(border_style="medium", color="000000")
            for col in range(min_col, max_col + 1):
                ws.cell(row=min_row, column=col).border = Border(top=thick)
            for col in range(min_col, max_col + 1):
                ws.cell(row=max_row, column=col).border = Border(bottom=thick)
            for row in range(min_row, max_row + 1):
                ws.cell(row=row, column=min_col).border = Border(left=thick)
            for row in range(min_row, max_row + 1):
                ws.cell(row=row, column=max_col).border = Border(right=thick)
            ws.cell(row=min_row, column=min_col).border = Border(top=thick, left=thick)
            ws.cell(row=min_row, column=max_col).border = Border(top=thick, right=thick)
            ws.cell(row=max_row, column=min_col).border = Border(bottom=thick, left=thick)
            ws.cell(row=max_row, column=max_col).border = Border(bottom=thick, right=thick)

        def thick_boder_table2(ws, min_row, max_row, min_col, max_col):
            thick = Side(border_style="medium", color="000000")
            thin = Side(border_style="thin", color="000000")

            for col in range(min_col, max_col + 1):
                ws.cell(row=min_row, column=col).border = Border(top=thick, bottom=thin, left=thin, right=thin)
            for col in range(min_col, max_col + 1):
                ws.cell(row=max_row, column=col).border = Border(bottom=thick, top=thin, left=thin, right=thin)
            for row in range(min_row, max_row + 1):
                ws.cell(row=row, column=min_col).border = Border(left=thick, right=thin, top=thin, bottom=thin)
            for row in range(min_row, max_row + 1):
                ws.cell(row=row, column=max_col).border = Border(right=thick, top=thin, left=thin, bottom=thin)

            ws.cell(row=min_row, column=min_col).border = Border(top=thick, left=thick, right=thin, bottom=thin)
            ws.cell(row=min_row, column=max_col).border = Border(top=thick, right=thick, bottom=thin, left=thin)
            ws.cell(row=max_row, column=min_col).border = Border(bottom=thick, left=thick, top=thin, right=thin)
            ws.cell(row=max_row, column=max_col).border = Border(bottom=thick, right=thick, left=thin, top=thin)

        thick_boder_table(worksheet, min_row=1, max_row=1, min_col=1, max_col=col_end)
        for col in range(3, col_end, 5):
            thick_boder_table(worksheet, min_row=2, max_row=5, min_col=col, max_col=col_end)
        for col in range(3, col_end, 5):
            thick_boder_table2(worksheet, min_row=6, max_row=max_row, min_col=col, max_col=col_end)

        worksheet.freeze_panes = 'C9'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        workbook.save(response)

        return response
    else:
        return HttpResponse(status=204)

@login_required()
def product_info_report(request):
    today = datetime.now().strftime('%Y-%m-%d')
    plants = Plant.objects.all()
    if request.method == 'POST':
        request.session['plant'] = request.POST.get('plant', '')
        request.session['mach'] = request.POST.get('mach', '')
        request.session['data_date'] = request.POST.get('data_date', '')
        request.session['to_date'] = request.POST.get('to_date', '')

    sPlant = request.session.get('plant', '')
    sMach = request.session.get('mach', '')
    sData_date = request.session.get('data_date', '')
    sTo_date = request.session.get('to_date', '')
    machs = Machine.objects.filter(plant=sPlant) if sPlant else None
    try:
        base_sql = """
            SELECT data_date, plant_id, mach_id, coagulant_time_hour, coagulant_time_min,
                   latex_time_hour, latex_time_min, tooling_time_hour, tooling_time_min,
                   remark, remark2, create_at, update_at
            FROM [VNEDC].[dbo].[collection_daily_prod_info]
        """
        conditions = []
        if sPlant:
            conditions.append(f"plant_id = '{sPlant}'")
        if sMach:
            conditions.append(f"mach_id = '{sMach}'")
        if sData_date and sTo_date:
            date_condition = f"CONVERT(DATE, create_at) BETWEEN CONVERT(DATE, '{sData_date}') AND CONVERT(DATE, '{sTo_date}')"
        elif sData_date:
            date_condition = f"CONVERT(DATE, create_at) BETWEEN CONVERT(DATE, '{sData_date}') AND DATEADD(DAY, 30, CONVERT(DATE, '{sData_date}'))"
        elif sTo_date:
            date_condition = f"CONVERT(DATE, create_at) BETWEEN DATEADD(DAY, -30, CONVERT(DATE, '{sTo_date}')) AND CONVERT(DATE, '{sTo_date}')"
        else:
            date_condition = f"CONVERT(DATE, create_at) BETWEEN DATEADD(DAY, -30, CONVERT(DATE, '{today}')) AND CONVERT(DATE, '{today}')"
        conditions.append(date_condition)
        sql = f"{base_sql} WHERE {' AND '.join(conditions)} ORDER BY data_date DESC, mach_id"
        db = vnedc_database()
        results = db.select_sql_dict(sql)
    except Exception as e:
        print(e)
        pass
    return render(request, 'collection/product_info_report.html', locals())


def oee_report(request):
    try:
        main_df_list = []
        line_data = []
        plants = Plant.objects.all()
        machs = Machine.objects.none()
        today_work_date = (datetime.now() - timedelta(days=1) - timedelta(hours=5)).strftime('%Y-%m-%d')
        # Retrieve session variables or set defaults
        sPlant = request.session.get('plant', '')
        sMach = request.session.get('mach', '')
        sData_date = request.session.get('data_date', datetime.today().strftime("%Y-%m-%d"))
        period_times = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 0, 1, 2, 3, 4, 5]
        sLimit = request.session.get('number1', 10)
        stopLimit = request.session.get('number2', 15)

        if request.method == 'POST':
            form_type = request.POST.get('form_type')
            if form_type == 'method_1':
                sData_date = request.POST.get('data_date', datetime.today().strftime("%Y-%m-%d"))
                sPlant = request.POST.get('plant')
                sMach = request.POST.get('mach')
                sLimit = int(request.POST.get('number1', 10))
                stopLimit = int(request.POST.get('number2', 15))
                request.session['data_date'] = sData_date
                request.session['plant'] = sPlant
                request.session['mach'] = sMach
                request.session['number1'] = sLimit
                request.session['number2'] = stopLimit
            elif form_type == 'method_2':
                sLimit = 10 if request.POST.get('number1', 10) == '' else int(request.POST.get('number1', 10))
                request.session['number1'] = sLimit
            elif form_type == 'method_3':
                stopLimit = 0 if request.POST.get('number2', 15) == '' else int(request.POST.get('number2', 15))
                request.session['number2'] = stopLimit
        if request.is_ajax():
            sData_date = request.GET.get('data_date')
            sPlant = request.GET.get('plant')
            sMach = request.GET.get('mach', datetime.today().strftime("%Y-%m-%d"))
            sButton = request.GET.get('button', '')
            sLimit = int(request.GET.get('limit', 10))
            stopLimit = int(request.GET.get('stop', 15))
        vnedc_db = vnedc_database()
        mes_db = mes_database()
        if sPlant == 'LK':
            sPlant = 'GDNBR'
            sPlant2 = 'NBR'
            up_limit = 'UpperLineSpeed_Min'
            low_limit = 'LowerLineSpeed_Min'
        elif sPlant == 'GDPVC':
            sPlant2 = 'PVC'
            up_limit = 'UpperSpeed'
            low_limit = 'LowerSpeed'
        elif sPlant == 'GDNBR':
            sPlant2 = 'NBR'
            up_limit = 'UpperLineSpeed_Min'
            low_limit = 'LowerLineSpeed_Min'
        machs = Machine.objects.filter(plant=sPlant)

        if sPlant != '' and sMach != '' and sPlant in sMach:
            # try:
            split_sMach = [sMach[:2], sMach[2:5], sMach[5:]]
            sql = f"""
                        SELECT distinct(cdm.Line), dml.name
                        FROM [PMGMES].[dbo].[PMG_DML_DataModelList] dml 
                        join [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] cdm
                        on cdm.MES_MACHINE = dml.name
                        where dml.DataModelTypeId= 'DMT000003' 
                        and dml.name like '%{split_sMach[0]}%' and dml.name like '%{split_sMach[1]}%' and dml.name like '%{split_sMach[2]}%'
                        order by cdm.line"""
            mach_list = mes_db.select_sql_dict(sql)
            mach = mach_list[0]['name']
            line_list = [item['Line'] for item in mach_list]

            sql1 = f"""WITH ConsecutiveStops AS (
                            SELECT MES_MACHINE, MachineName, LINE, CAST(DATEPART(hour, CreationTime) AS INT) AS Period, Qty2 AS Qty, CreationTime AS Cdt,
                                CASE WHEN Qty2 IS NULL OR Qty2 <= {sLimit} THEN 1 ELSE 0 END AS IsStop, --StopLimitHere!
                                ROW_NUMBER() OVER (PARTITION BY MES_MACHINE, LINE ORDER BY CreationTime)
                                - ROW_NUMBER() OVER (PARTITION BY MES_MACHINE, LINE, CASE WHEN Qty2 IS NULL OR Qty2 <= {sLimit} THEN 1 ELSE 0 END ORDER BY CreationTime) AS StopGroup  --StopLimitHere!
                            FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] d
                            JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m
                                ON d.MachineName = m.COUNTING_MACHINE
                            WHERE m.MES_MACHINE = '{mach}' AND CreationTime BETWEEN CONVERT(DATETIME, '{sData_date} 06:00:00', 120) AND DATEADD(DAY, 1, CONVERT(DATETIME, '{sData_date} 05:59:59', 120))),
                        GroupStats AS (
                            SELECT MES_MACHINE, LINE, StopGroup, COUNT(*) AS StopRowCount, MIN(Cdt) AS StartCdt, MAX(Cdt) AS EndCdt
                            FROM ConsecutiveStops
                            WHERE IsStop = 1
                            GROUP BY MES_MACHINE, LINE, StopGroup)
                        SELECT cs.MES_MACHINE as MES_MachineName, cs.MachineName, cs.LINE, cast(cs.Period as int) as Period, cast(cs.Qty as int) as Quantity, cs.Cdt as CreationTime, CASE WHEN gs.StopRowCount > {int(stopLimit/5)} THEN 5 ELSE 0 END AS Stop_time
                        FROM ConsecutiveStops cs
                        LEFT JOIN GroupStats gs
                        ON cs.MES_MACHINE = gs.MES_MACHINE AND cs.LINE = gs.LINE AND cs.Cdt BETWEEN gs.StartCdt AND gs.EndCdt
                        ORDER BY cs.MES_MACHINE, cs.LINE, cs.Cdt;"""
            detail_raws = mes_db.select_sql_dict(sql1)
            stop_time_df = pd.DataFrame(detail_raws)
            stop_time_df = stop_time_df.fillna('')
            stop_time_df['Period'] = stop_time_df['Period'].astype(str)

            sap_sql = f"""
                    SELECT wo.Id WorkOrder, wo.PartNo, ProductItem, rc.InspectionDate, rc.LineName Line, rc.Id Runcard, Period, CustomerName, AQL, PackingType, ActualQty SAPQuantity
                    FROM [dbo].[PMG_MES_WorkInProcess] wi
                    JOIN [PMGMES].[dbo].[PMG_MES_RunCard] rc on rc.Id = wi.RunCardId
                    JOIN [PMGMES].[dbo].[PMG_MES_WorkOrder] wo on wi.WorkOrderId = wo.Id
                    WHERE rc.MachineName = '{mach}'
                    AND ((rc.InspectionDate = '{sData_date}' AND rc.Period BETWEEN 6 AND 23)
                    OR (rc.InspectionDate = DATEADD(DAY, 1, '{sData_date}') AND rc.Period BETWEEN 0 AND 5))
                    ORDER BY rc.LineName, rc.InspectionDate, CAST(rc.Period AS INT)
                    """
            sap_raw = mes_db.select_sql_dict(sap_sql)
            sap_raw_df = pd.DataFrame(sap_raw)

            sql = f"""
                    WITH COUNTING_DATA AS (
                                SELECT MES_MACHINE, LINE, CAST(DATEPART(hour, CreationTime) AS INT) Period, SUM(case when Qty2 >= 0 then Qty2 else 0 end) Qty,
                                    DATEADD(HOUR, DATEDIFF(HOUR, 0, CreationTime), 0) Cdt
                                FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] d
                                JOIN [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m
                                    ON d.MachineName = m.COUNTING_MACHINE
                                WHERE m.MES_MACHINE = '{mach}' AND CreationTime BETWEEN CONVERT(DATETIME, '{sData_date} 06:00:00', 120)
                                    AND DATEADD(DAY, 1, CONVERT(DATETIME, '{sData_date} 05:59:59', 120))
                                GROUP BY m.MES_MACHINE, LINE, CAST(DATEPART(hour, CreationTime) AS INT), DATEADD(HOUR, DATEDIFF(HOUR, 0, CreationTime), 0)
                            )
                            SELECT
                                wo.Id AS WorkOrder, wo.PartNo, wo.ProductItem, rc.InspectionDate AS InspectDate, rc.MachineName AS Machine, rc.LineName AS Line,
                                rc.Id as Runcard, rc.period AS Period,  cd.Qty as CoutingQuantity, std.{low_limit} AS LowSpeed, std.{up_limit} AS UpSpeed,
                                CAST(ROUND((std.{low_limit} + std.{up_limit}) / 2, 1) AS FLOAT) AS StdSpeed,
                                SUM(CASE WHEN ft.ActualQty IS NOT NULL THEN ft.ActualQty ELSE 0 END) AS FaultyQuantity,
                                SUM(CASE WHEN sp.ActualQty IS NOT NULL THEN sp.ActualQty ELSE 0 END) AS ScrapQuantity,
                                MAX(ipqc.InspectionValue) AS InspectionValue, MIN(wo.StartDate) AS StartDate, MAX(wo.EndDate) AS EndDate
                            FROM [PMGMES].[dbo].[PMG_MES_RunCard] rc
                            JOIN [PMGMES].[dbo].[PMG_MES_WorkOrder] wo
                                ON wo.id = rc.WorkOrderId AND wo.StartDate IS NOT NULL
                            LEFT JOIN [PMGMES].[dbo].[PMG_MES_{sPlant2}_SCADA_Std] std
                                ON std.PartNo = wo.PartNo
                            LEFT JOIN [PMGMES].[dbo].[PMG_MES_Faulty] ft
                                ON ft.RunCardId = rc.id AND ft.WorkOrderId = wo.id
                            LEFT JOIN [PMGMES].[dbo].[PMG_MES_Scrap] sp
                                ON sp.RunCardId = rc.id AND sp.WorkOrderId = wo.id
                            LEFT JOIN COUNTING_DATA cd
                                ON cd.MES_MACHINE = rc.MachineName AND rc.LineName = cd.LINE AND rc.Period = cd.Period
                            LEFT JOIN [PMGMES].[dbo].[PMG_MES_IPQCInspectingRecord] ipqc
                                ON ipqc.RunCardId = rc.Id AND ipqc.OptionName = 'Weight'
                            WHERE rc.MachineName = '{mach}'
                                AND ((rc.InspectionDate = '{sData_date}' AND rc.Period BETWEEN 6 AND 23)
                                    OR (rc.InspectionDate = DATEADD(DAY, 1, '{sData_date}') AND rc.Period BETWEEN 0 AND 5))
                                --AND cd.Cdt > wo.StartDate
                                AND (cd.Cdt < wo.EndDate OR wo.EndDate IS NULL)
                            GROUP BY wo.Id, wo.PartNo,
                                wo.ProductItem, rc.InspectionDate, rc.MachineName, rc.LineName, rc.Id, rc.period, cd.Qty,
                                std.{low_limit}, std.{up_limit}
                            ORDER BY rc.LineName, rc.InspectionDate, CAST(rc.Period AS INT)

                    """
            detail_raws = mes_db.select_sql_dict(sql)
            len_detail_raws = len(detail_raws)

            if len_detail_raws > 0:
                raw_df = pd.DataFrame(detail_raws)
                # raw_df['Stop_time'] = raw_df['Stop_time'].apply(lambda x: 0 if x <= stopLimit else x)
                for line in line_list:
                    main_df_list_item = {}
                    data = {}
                    data_df = raw_df[raw_df['Line'] == line]

                    data_df = data_df.copy()  # Create an independent copy
                    data_df.loc[:, 'Period'] = data_df['Period'].astype(int)

                    # Same Period 2 Runcard, check which one is real
                    duplicated_periods = data_df[data_df.duplicated(subset=['Period'], keep=False)]
                    if not duplicated_periods.empty:
                        runcards_to_check = duplicated_periods['Runcard'].unique()
                        sql = f"""
                        SELECT DISTINCT RunCardId
                        FROM [PMGMES].[dbo].[PMG_MES_IPQCInspectingRecord]
                        WHERE RunCardId IN ({','.join(f"'{r}'" for r in runcards_to_check)})
                          AND OptionName = 'Weight'
                        """
                        ipqc_list = mes_db.select_sql_dict(sql)
                        ipqc_list = [item['RunCardId'] for item in ipqc_list]
                        # Step 4: Filter duplicated Periods by IPQC_List
                        filtered_duplicates = duplicated_periods[duplicated_periods['Runcard'].isin(ipqc_list)]
                        final_df = pd.concat(
                            [data_df[~data_df['Period'].isin(duplicated_periods['Period'])], filtered_duplicates])
                        data_df = final_df.sort_values(by=['Line', 'InspectDate', 'Period']).reset_index(drop=True)

                    line_stop_time_df = stop_time_df[stop_time_df['LINE'] == line].copy()

                    line_stop_time_df.loc[:, 'Quantity'] = pd.to_numeric(line_stop_time_df['Quantity'], errors='coerce')

                    stop_time_df_filtered_df = line_stop_time_df[(line_stop_time_df['Stop_time'] == 5)]

                    machine_stop_time = len(stop_time_df_filtered_df) * 5
                    activation = round((1440 - machine_stop_time) / 1440, 2) if machine_stop_time > 0 else 1
                    minus_machine_stop_time = 1440 - machine_stop_time

                    stop_time_df_filtered_df = stop_time_df_filtered_df.copy()
                    stop_time_df_filtered_df.loc[len(stop_time_df_filtered_df)] = {'MES_MachineName': '',
                                                                                   'MachineName': '', 'LINE': '',
                                                                                   'Period': '', 'Quantity': '',
                                                                                   'CreationTime': '-Total-',
                                                                                   'Stop_time': line_stop_time_df[
                                                                                       'Stop_time'].apply(
                                                                                       lambda x: int(x) if str(
                                                                                           x).isdigit() else 0).sum()}

                    activation = round((1440 - machine_stop_time) / 1440, 3) if machine_stop_time > 0 else 1
                    activation_text = f'{round(activation*100,1)}%'

                    sstop_table = []

                    line_stop_time_df.loc[:, 'CreationTime'] = pd.to_datetime(line_stop_time_df['CreationTime'])
                    line_stop_time_df.loc[:, 'Period'] = line_stop_time_df['CreationTime'].dt.hour
                    line_stop_time_df.loc[:, 'Stop_time'] = line_stop_time_df['Stop_time'].replace('', 0).astype(int)
                    stop_time_summary = line_stop_time_df.groupby('Period')['Stop_time'].sum().reset_index()

                    period_str_values = stop_time_summary['Period'].values

                    for period in period_times:
                        if period in stop_time_summary['Period'].values:
                            stop_time_value = \
                            stop_time_summary.loc[stop_time_summary['Period'] == period, 'Stop_time'].iloc[0]
                            sstop_table.append(stop_time_value)
                        else:
                            sstop_table.append(100)

                    stop_table = [period_times[:12], sstop_table[:12], period_times[12:], sstop_table[12:]]

                    stop_time_summary['Period'] = stop_time_summary['Period'].astype(int)
                    data_df = pd.merge(data_df, stop_time_summary, on=['Period'], how='left')
                    data_df['Run_time'] = 60 - data_df['Stop_time']
                    data_df['Target'] = data_df['Run_time'] * data_df['StdSpeed']

                    speed_df = data_df[
                        ['WorkOrder', 'PartNo', 'ProductItem', 'InspectDate', 'Machine', 'Line', 'Runcard', 'Period',
                         'LowSpeed', 'UpSpeed', 'StdSpeed', 'Run_time', 'Target']].copy()
                    speed_df = speed_df.drop_duplicates(subset=[
                        'WorkOrder', 'PartNo', 'ProductItem', 'InspectDate',
                        'Machine', 'Line', 'Runcard', 'Period',
                        'LowSpeed', 'UpSpeed', 'StdSpeed', 'Run_time', 'Target'
                    ])

                    # speed = speed_df[speed_df['StdSpeed'].notna()]['StdSpeed'].sum()/len(speed_df[speed_df['StdSpeed'].notna()]['StdSpeed'])
                    speed = speed_df['StdSpeed'].dropna().mean()
                    speed_df.loc[len(speed_df)] = {'WorkOrder': '', 'PartNo': '', 'ProductItem': '', 'InspectDate': '',
                                                   'Machine': '', 'Line': '', 'Runcard': '', 'Period': '',
                                                   'LowSpeed': '', 'UpSpeed': '', 'StdSpeed': '', 'Run_time': '-Total-',
                                                   'Target': speed_df['Target'].sum()}
                    run_time = 1440 - machine_stop_time
                    estimate_ouput = int(run_time * speed)

                    counting_df = data_df[
                        ['WorkOrder', 'PartNo', 'ProductItem', 'InspectDate', 'Machine', 'Line', 'Runcard', 'Period',
                         'CoutingQuantity', 'FaultyQuantity', 'ScrapQuantity']].copy()
                    counting_df = counting_df.drop_duplicates(subset=[
                        'WorkOrder', 'PartNo', 'ProductItem', 'InspectDate',
                        'Machine', 'Line', 'Runcard', 'Period',
                        'CoutingQuantity', 'FaultyQuantity', 'ScrapQuantity'
                    ])
                    couting_data = int(counting_df['CoutingQuantity'].sum())

                    second_data = int(counting_df['FaultyQuantity'].sum())

                    scrap_data = int(counting_df['ScrapQuantity'].sum())

                    counting_df = pd.concat([counting_df, pd.DataFrame([{'WorkOrder': '', 'PartNo': '',
                                                                         'ProductItem': '', 'InspectDate': '',
                                                                         'Machine': '', 'Line': '',
                                                                         'Runcard': '', 'Period': '-Total-',
                                                                         'CoutingQuantity': couting_data,
                                                                         'FaultyQuantity': second_data,
                                                                         'ScrapQuantity': scrap_data}])],
                                            ignore_index=True)
                    capacity = 0 if estimate_ouput == 0 else round(
                        (couting_data + second_data + scrap_data) / estimate_ouput, 3)
                    capacity_text = f'{round(capacity*100,1)}%'

                    sap_df = sap_raw_df[sap_raw_df['Line'] == line]
                    sap_data = int(sap_df['SAPQuantity'].sum())
                    sap_df = sap_df.copy()
                    sap_df = pd.concat([sap_df, pd.DataFrame([{'WorkOrder': '', 'PartNo': '',
                                                               'ProductItem': '', 'InspectionDate': '',
                                                               'Line': '', 'Runcard': '', 'Period': '',
                                                               'CustomerName': '', 'AQL': '',
                                                               'PackingType': '-Total-', 'SAPQuantity': sap_data}])],
                                       ignore_index=True)

                    yield_data = round(sap_data / (couting_data + second_data + scrap_data), 3)
                    yield_data_text = f'{round(yield_data*100,1)}%'

                    oee_data = round(activation * capacity * yield_data, 3)
                    oee_data_text = f'{round(oee_data*100,1)}%'

                    main_df_list_item['line_name'] = line
                    main_df_list_item['stop_time_df'] = stop_time_df_filtered_df
                    main_df_list_item['speed_df'] = speed_df
                    main_df_list_item['counting_df'] = counting_df
                    main_df_list_item['second_df'] = counting_df
                    main_df_list_item['scrap_df'] = counting_df
                    main_df_list_item['sap_df'] = sap_df
                    main_df_list.append(main_df_list_item)

                    data["line_name"] = line
                    data["machine_stop_time"] = machine_stop_time
                    data["stop_table"] = stop_table
                    data["activation"] = activation
                    data["activation_text"] = activation_text
                    data["speed"] = speed
                    data["run_time"] = run_time
                    data["couting_data"] = couting_data
                    data["second_data"] = second_data
                    data["scrap_data"] = scrap_data
                    data["estimate_ouput"] = estimate_ouput
                    data["capacity"] = capacity_text
                    data["sap_data"] = sap_data
                    data["yield_data"] = yield_data_text
                    data["oee_data"] = oee_data_text
                    line_data.append(data)

                if request.is_ajax() or (request.method == 'POST' and form_type == 'method_2'):
                    try:
                        if 'id-dom' not in sButton:
                            for line in line_list:
                                if line in sButton:
                                    if 'id-machine-stop-time' in sButton:
                                        table_html = next((item['stop_time_df'] for item in main_df_list if
                                                           item['line_name'] == f'{line}'), None).to_html(
                                            classes='table table-striped', index=False)
                                    elif 'id-speed' in sButton:
                                        table_html = next((item['speed_df'] for item in main_df_list if
                                                           item['line_name'] == f'{line}'), None).to_html(
                                            classes='table table-striped', index=False)
                                    elif 'id-counting-machine-qty' in sButton:
                                        table_html = next((item['counting_df'] for item in main_df_list if
                                                           item['line_name'] == f'{line}'), None).to_html(
                                            classes='table table-striped', index=False)
                                    elif 'id-second-hand-qty' in sButton:
                                        table_html = next((item['second_df'] for item in main_df_list if
                                                           item['line_name'] == f'{line}'), None).to_html(
                                            classes='table table-striped', index=False)
                                    elif 'id-scrap-qty' in sButton:
                                        table_html = next((item['scrap_df'] for item in main_df_list if
                                                           item['line_name'] == f'{line}'), None).to_html(
                                            classes='table table-striped', index=False)
                                    elif 'id-sap-qty' in sButton:
                                        table_html = next(
                                            (item['sap_df'] for item in main_df_list if item['line_name'] == f'{line}'),
                                            None).to_html(classes='table table-striped', index=False)
                                    return JsonResponse({'table': table_html}, safe=False)
                    except Exception as e:
                        print(e)
                        pass
                parameters = ['稼動率', '預估產能', '產能效率', '良率', 'OEE']
            # for line, data in zip(line_list, line_data):
            #     print(f"{line}: {data}")
    except Exception as e:
        print(e)
    return render(request, 'collection/oee_report.html', locals())


# DATA_IMPORT

def to_int_or_none(value):
    """Chuyển đổi giá trị thành int nếu không phải NaN, ngược lại trả về None"""
    return int(value) if pd.notna(value) else None


def to_float_or_none(value):
    return float(value) if pd.notna(value) else None


def to_boolean_or_none(value):
    if pd.isna(value):  # Nếu là NaN thì trả về None
        return None

    if isinstance(value, (int, float)):  # Nếu là số
        return bool(value)  # 0 -> False, 1 -> True

    str_value = str(value).strip().lower()  # Chuyển đổi thành chuỗi, loại bỏ khoảng trắng

    if str_value in ["true", "1", "yes"]:
        return True
    elif str_value in ["false", "0", "no"]:
        return False

    return None  # Nếu không khớp, trả về None


def to_string_or_none(value):
    return "" if pd.isna(value) else str(value)


def to_string_or_null(value):
    return None if pd.isna(value) else str(value)


def plant_sheet(df, request):
    plant_codes = df["Plant Code"].unique()
    plant_dict = {p.plant_code: p for p in Plant.objects.filter(plant_code__in=plant_codes)}

    plants_to_update = []
    plants_to_create = []

    for _, row in df.iterrows():
        plant_code = to_string_or_none(row["Plant Code"])
        plant_name = to_string_or_none(row["Plant Name"])

        if plant_code in plant_dict:
            obj = plant_dict[plant_code]
            if obj.plant_code != plant_code or obj.plant_name != plant_name:
                obj.plant_code = plant_code
                obj.plant_name = plant_name
                obj.update_at = timezone.now()
                obj.update_by = request.user
                plants_to_update.append(obj)
        else:
            plants_to_create.append(Plant(
                plant_code=plant_code,
                plant_name=plant_name,
                update_at=timezone.now(),
                update_by=request.user
            ))

    with transaction.atomic():
        if plants_to_update:
            Plant.objects.bulk_update(plants_to_update, ["plant_name", "update_at", "update_by"])
        if plants_to_create:
            Plant.objects.bulk_create(plants_to_create)


def machine_sheet(df, request):
    mach_codes = df["Mach Code"].unique()
    mach_dict = {m.mach_code: m for m in Machine.objects.filter(mach_code__in=mach_codes)}

    machines_to_update = []
    machines_to_create = []

    for _, row in df.iterrows():
        mach_code = to_string_or_none(row["Mach Code"])
        mach_name = to_string_or_none(row["Mach Name"])
        mold_type = to_string_or_none(row["Mold Type"])
        plant_code = to_string_or_none(row["Plant"])

        if mach_code in mach_dict:
            obj = mach_dict[mach_code]
            if obj.mach_code != mach_code or obj.mach_name != mach_name or obj.mold_type != mold_type \
                    or str(obj.plant) != plant_code:
                obj.mach_code = mach_code
                obj.mach_name = mach_name
                obj.mold_type = mold_type
                obj.plant = Plant.objects.get(plant_code=plant_code)
                obj.update_at = timezone.now()
                obj.update_by = request.user
                machines_to_update.append(obj)
        else:
            plant_instance = Plant.objects.filter(plant_code=plant_code).first()

            machines_to_create.append(Machine(
                mach_code=mach_code,
                mach_name=mach_name,
                mold_type=mold_type,
                plant=plant_instance,
                update_at=timezone.now(),
                update_by=request.user
            ))

    with transaction.atomic():
        if machines_to_update:
            Machine.objects.bulk_update(machines_to_update, ["mach_name", "mold_type", "plant",
                                                             "update_at", "update_by"])
        if machines_to_create:
            Machine.objects.bulk_create(machines_to_create)


def process_type_sheet(df, request):
    process_type_codes = df["Process Code"].unique()
    process_type_dict = {p.process_code: p for p in Process_Type.objects.filter(process_code__in=process_type_codes)}

    process_types_to_update = []
    process_types_to_create = []

    for _, row in df.iterrows():
        process_code = to_string_or_none(row["Process Code"])
        process_name = to_string_or_none(row["Process Name"])
        process_tw = to_string_or_none(row["Traditional Chinese Name"])
        process_cn = to_string_or_none(row["Simple Chinese Name"])
        process_vn = to_string_or_none(row["Vietnamese Name"])
        show_order = to_int_or_none(row["Order"])

        if process_code in process_type_dict:
            obj = process_type_dict[process_code]
            if obj.process_code != process_code or obj.process_name != process_name or obj.process_tw != process_tw \
                    or obj.process_cn != process_cn or obj.process_vn != process_vn or obj.show_order != show_order:
                obj.process_code = process_code
                obj.process_name = process_name
                obj.process_tw = process_tw
                obj.process_cn = process_cn
                obj.process_vn = process_vn
                obj.show_order = show_order
                obj.update_at = timezone.now()
                obj.update_by = request.user
                process_types_to_update.append(obj)
        else:
            process_types_to_create.append(Process_Type(
                process_code=process_code,
                process_name=process_name,
                process_tw=process_tw,
                process_cn=process_cn,
                process_vn=process_vn,
                show_order=show_order,
                update_at=timezone.now(),
                update_by=request.user
            ))

    with transaction.atomic():
        if process_types_to_update:
            Process_Type.objects.bulk_update(process_types_to_update, ["process_name", "process_tw",
                                                                       "process_cn", "process_vn", "show_order",
                                                                       "update_at", "update_by"])
        if process_types_to_create:
            Process_Type.objects.bulk_create(process_types_to_create)


def parameter_type_sheet(df, request):
    param_type_codes = df["Parameter Type Code"].unique()
    process_type_codes = df["Process Type"].unique()

    # Lấy danh sách Process_Type theo process_code
    process_type_dict = {
        p.process_code: p for p in Process_Type.objects.filter(process_code__in=process_type_codes)
    }

    # Tạo dictionary chứa Parameter_Type
    param_type_dict = {
        (p.param_type_code, p.process_type.process_code): p
        for p in Parameter_Type.objects.filter(
            param_type_code__in=param_type_codes,
            process_type__in=process_type_dict.values()
        )
    }

    param_types_to_update = []
    param_types_to_create = []

    for _, row in df.iterrows():
        param_type_code = to_string_or_none(row["Parameter Type Code"])
        process_type_code = to_string_or_none(row["Process Type"])
        control_table = to_string_or_none(row["Control Table"])
        control_high_column = to_string_or_none(row["Control High Column"])
        control_low_column = to_string_or_none(row["Control Low Column"])

        # Lấy object process_type từ dictionary
        process_type_ = process_type_dict.get(process_type_code)
        if not process_type_:
            continue  # Bỏ qua nếu không tìm thấy process_type

        key = (param_type_code, process_type_code)  # Dùng process_code thay vì id

        if key in param_type_dict:
            obj = param_type_dict[key]
            if (obj.control_table != control_table or
                    obj.control_high_column != control_high_column or
                    obj.control_low_column != control_low_column):
                obj.control_table = control_table
                obj.control_high_column = control_high_column
                obj.control_low_column = control_low_column
                obj.update_at = timezone.now()
                obj.update_by = request.user
                param_types_to_update.append(obj)
        else:
            param_types_to_create.append(Parameter_Type(
                param_type_code=param_type_code,
                process_type=process_type_,
                control_table=control_table,
                control_high_column=control_high_column,
                control_low_column=control_low_column,
                update_at=timezone.now(),
                update_by=request.user
            ))

    # Cập nhật dữ liệu vào database
    with transaction.atomic():
        if param_types_to_update:
            Parameter_Type.objects.bulk_update(param_types_to_update, [
                "control_table", "control_high_column", "control_low_column", "update_at", "update_by"
            ])
        if param_types_to_create:
            Parameter_Type.objects.bulk_create(param_types_to_create)


def param_define_sheet(df, request):
    plant_codes = df["Plant"].unique()
    machine_codes = df["Machine"].unique()
    process_type_codes = df["Process Type"].unique()
    param_type_codes = df["Param Type"].unique()
    param_name_codes = df["Parameter Name"].unique()

    plant_dict = {p.plant_code: p for p in Plant.objects.filter(plant_code__in=plant_codes)}
    machine_dict = {m.mach_code: m for m in Machine.objects.filter(mach_code__in=machine_codes)}
    process_type_dict = {p.process_code: p for p in Process_Type.objects.filter(process_code__in=process_type_codes)}

    param_define_dict = {
        (p.plant.plant_code, p.mach.mach_code, p.process_type.process_code, p.param_type, p.parameter_name): p
        for p in ParameterDefine.objects.filter(
            plant__in=plant_dict.values(),
            mach__in=machine_dict.values(),
            process_type__in=process_type_dict.values(),
            param_type__in=param_type_codes,
            parameter_name__in=param_name_codes
        )
    }

    param_define_to_update = []
    param_define_to_create = []

    for _, row in df.iterrows():
        plant_code = to_string_or_none(row["Plant"])
        machine_code = to_string_or_none(row["Machine"])
        process_type_code = to_string_or_none(row["Process Type"])
        param_type_code = to_string_or_none(row["Param Type"])
        parameter_name_code = to_string_or_none(row["Parameter Name"])

        if not all([plant_code, machine_code, process_type_code, param_type_code, parameter_name_code]):
            continue

        parameter_tw = to_string_or_null(row["Traditional Chinese Name"])
        parameter_cn = to_string_or_null(row["Simple Chinese Name"])
        parameter_vn = to_string_or_null(row["Vietnamese Name"])

        show_order = to_int_or_none(row["Order"])
        base_line = to_float_or_none(row["Base Line"])
        control_range_high = to_float_or_none(row["Control Range High"])
        control_range_low = to_float_or_none(row["Control Range Low"])
        sampling_frequency = to_string_or_null(row["Sampling Frequency"])
        unit = to_string_or_null(row["Unit"])
        auto_value = to_boolean_or_none(row["Auto Value"])
        pda_value = to_boolean_or_none(row["PDA Value"])
        side = to_string_or_null(row["Side"])
        scada_column = to_string_or_null(row["Scada Column"])
        scada_table = to_string_or_null(row["Scada Table"])
        input_time = to_string_or_null(row["Input Time"])
        text_color = to_string_or_null(row["Text Color"])

        plant = plant_dict.get(plant_code)
        machine = machine_dict.get(machine_code)
        process_type = process_type_dict.get(process_type_code)

        if not plant or not machine or not process_type:
            continue

        key = (plant_code, machine_code, process_type_code, param_type_code, parameter_name_code)

        if key in param_define_dict:
            obj = param_define_dict[key]
            if (obj.parameter_tw != parameter_tw or obj.parameter_cn != parameter_cn or
                    obj.parameter_vn != parameter_vn or obj.show_order != show_order or obj.base_line != base_line or
                    obj.control_range_high != control_range_high or obj.control_range_low != control_range_low or
                    obj.sampling_frequency != sampling_frequency or obj.unit != unit or obj.auto_value != auto_value or
                    obj.pda_value != pda_value or obj.side != side or obj.scada_column != scada_column or
                    obj.scada_table != scada_table or obj.input_time != input_time or obj.text_color != text_color):
                obj.parameter_tw = parameter_tw
                obj.parameter_cn = parameter_cn
                obj.parameter_vn = parameter_vn
                obj.show_order = show_order
                obj.base_line = base_line
                obj.control_range_high = control_range_high
                obj.control_range_low = control_range_low
                obj.sampling_frequency = sampling_frequency
                obj.unit = unit
                obj.auto_value = auto_value
                obj.pda_value = pda_value
                obj.side = side
                obj.scada_column = scada_column
                obj.scada_table = scada_table
                obj.input_time = input_time
                obj.text_color = text_color
                obj.update_at = timezone.now()
                obj.update_by = request.user

                param_define_to_update.append(obj)
        else:
            param_define_to_create.append(ParameterDefine(
                plant=plant,
                mach=machine,
                process_type=process_type,
                param_type=param_type_code,
                side=side,
                parameter_name=parameter_name_code,
                parameter_tw=parameter_tw,
                parameter_cn=parameter_cn,
                parameter_vn=parameter_vn,
                show_order=show_order,
                unit=unit,
                control_range_low=control_range_low,
                base_line=base_line,
                control_range_high=control_range_high,
                sampling_frequency=sampling_frequency,
                input_time=input_time,
                auto_value=auto_value,
                scada_table=scada_table,
                scada_column=scada_column,
                text_color=text_color,
                pda_value=pda_value,
                create_at=timezone.now(),
                create_by=request.user
            ))

    with transaction.atomic():
        if param_define_to_update:
            ParameterDefine.objects.bulk_update(param_define_to_update, [
                "parameter_cn", "parameter_vn", "show_order", "base_line", "control_range_high", "control_range_low",
                "sampling_frequency", "unit", "auto_value", "pda_value", "side", "scada_column", "scada_table",
                "input_time", "text_color", "update_at", "update_by", "parameter_tw"
            ])
        if param_define_to_create:
            ParameterDefine.objects.bulk_create(param_define_to_create)


def lab_parameter_sheet(df, request):
    plant_codes = df["Plant"].unique()
    machine_codes = df["Machine"].unique()
    item_no_codes = df["Item No"].unique()
    process_type_codes = df["Process Type"].unique()
    parameter_name_codes = df["Parameter Name"].unique()

    plant_dict = {p.plant_code: p for p in Plant.objects.filter(plant_code__in=plant_codes)}
    machine_dict = {m.mach_code: m for m in Machine.objects.filter(mach_code__in=machine_codes)}

    lab_param_dict = {
        (p.plant.plant_code, p.mach.mach_code, p.item_no, p.process_type, p.parameter_name): p
        for p in Lab_Parameter_Control.objects.filter(
            plant__in=plant_dict.values(),
            mach__in=machine_dict.values(),
            item_no__in=item_no_codes,
            process_type__in=process_type_codes,
            parameter_name__in=parameter_name_codes
        )
    }

    lab_params_to_update = []
    lab_params_to_create = []

    for _, row in df.iterrows():
        plant_code = to_string_or_none(row["Plant"])
        machine_code = to_string_or_none(row["Machine"])
        item_no = to_string_or_none(row["Item No"])
        process_type = to_string_or_none(row["Process Type"])
        parameter_name = to_string_or_none(row["Parameter Name"])

        if not all([plant_code, machine_code, item_no, process_type, parameter_name]):
            continue

        control_range_low = to_float_or_none(row["Control Range Low"])
        base_line = to_float_or_none(row["Base Line"])
        control_range_high = to_float_or_none(row["Control Range High"])

        plant = plant_dict.get(plant_code)
        machine = machine_dict.get(machine_code)

        if not plant or not machine:
            continue

        key = (plant_code, machine_code, item_no, process_type, parameter_name)

        if key in lab_param_dict:
            obj = lab_param_dict[key]
            if (obj.control_range_low != control_range_low or
                    obj.base_line != base_line or
                    obj.control_range_high != control_range_high):
                obj.control_range_low = control_range_low
                obj.base_line = base_line
                obj.control_range_high = control_range_high
                obj.create_at = timezone.now()
                obj.create_by = request.user
                lab_params_to_update.append(obj)
        else:
            lab_params_to_create.append(Lab_Parameter_Control(
                plant=plant,
                mach=machine,
                item_no=item_no,
                process_type=process_type,
                parameter_name=parameter_name,
                control_range_low=control_range_low,
                base_line=base_line,
                control_range_high=control_range_high,
                create_at=timezone.now(),
                create_by=request.user
            ))

    with transaction.atomic():
        if lab_params_to_update:
            Lab_Parameter_Control.objects.bulk_update(lab_params_to_update, [
                "control_range_low", "base_line", "control_range_high", "create_at", "create_by"
            ])
        if lab_params_to_create:
            Lab_Parameter_Control.objects.bulk_create(lab_params_to_create)


SHEET_PROCESSORS = {
    "Plant": plant_sheet,
    "Machine": machine_sheet,
    "Process Type": process_type_sheet,
    "Param Type": parameter_type_sheet,
    "Param Define": param_define_sheet,
    "Lab Param": lab_parameter_sheet
}


def import_excel_data(request):
    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            df_dict = pd.read_excel(excel_file, sheet_name=None)  # Đọc tất cả các sheet

            for sheet_name, df in df_dict.items():
                process_function = SHEET_PROCESSORS.get(sheet_name)
                if process_function:
                    process_function(df, request)  # Gọi function tương ứng

            return render(request, "collection/collection_data_import.html",
                          {"form": form, "message": "Upload successfully!"})

    else:
        form = ExcelUploadForm()

    return render(request, "collection/collection_data_import.html", locals())

# END_DATA_IMPORT