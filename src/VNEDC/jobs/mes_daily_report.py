import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from jobs.database import mes_database
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side
from openpyxl.styles import PatternFill


class mes_daily_report(object):
    report_date1 = ""
    report_date2 = ""

    def __init__(self, report_date1, report_date2):
        self.report_date1 = report_date1
        self.report_date2 = report_date2

    def send_email(self):
        # 配置发件人和收件人信息
        sender_email = "box.chang@egvnco.com"
        receiver_email = "phil.wang@egvnco.com"
        subject = "TEST"
        body = "TEST."

        # 创建一个带附件的MIMEMultipart对象
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = receiver_email
        msg['Subject'] = subject

        # 添加邮件正文
        msg.attach(MIMEText(body, 'plain'))

        # 添加附件（如果有）
        filename = "template.xlsx"  # 替换为你的文件名
        attachment = open(filename, "rb")

        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {filename}")

        msg.attach(part)

        # 发送邮件
        try:
            with smtplib.SMTP('mail.egvnco.com', 587) as server:  # 替换为你的SMTP服务器和端口
                server.starttls()  # 启用TLS（传输层安全）
                server.login(sender_email, "1qazxsw2")  # 使用你的邮箱和密码登录
                text = msg.as_string()
                server.sendmail(sender_email, receiver_email, text)
            print("Email sent successfully!")
        except Exception as e:
            print(f"Failed to send email: {e}")
        finally:
            attachment.close()

    def generate_excel(self):
        report_date1 = self.report_date1
        report_date2 = self.report_date2
        db = mes_database()

        # 工單資料
        sql = f"""
          SELECT w.MachineId,Name,wi.LineId Line,CAST(r.Period as INT) Period,wi.StartDate, wi.EndDate, wi.WorkOrderId,WorkOrderDate,CustomerName,PartNo,ProductItem,w.AQL,w.PlanQty,wi.Qty,w.Status
          FROM [PMG_MES_WorkOrderInfo] wi, [PMG_MES_WorkOrder] w, [PMG_DML_DataModelList] dl,[PMG_MES_RunCard] r
          where wi.WorkOrderId = w.id and w.MachineId = dl.Id and r.WorkOrderId = w.Id and r.LineName = wi.LineId
          and wi.StartDate between CONVERT(DATETIME, '{report_date1} 05:30:00', 120) and CONVERT(DATETIME, '{report_date2} 05:29:59', 120)
          order by Name,CAST(r.Period as INT),wi.LineId
        """
        raws = db.select_sql_dict(sql)

        df_main = pd.DataFrame(raws)

        # 增加欄位班別
        df_main['Shift'] = df_main['Period'].apply(self.shift)

        # 點數機資料
        sql = f"""
        SELECT FORMAT(CreationTime, 'yyyy-MM-dd') AS Date,CAST(DATEPART(hour, CreationTime) as INT) Period ,m.mes_machine Name,m.line Line, max(Speed) max_speed,min(Speed) min_speed,round(avg(Speed),0) avg_speed,sum(Qty2) sum_qty
          FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] c, [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m
          where CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120)
          and c.MachineName = m.counting_machine 
          group by m.mes_machine,FORMAT(CreationTime, 'yyyy-MM-dd'),DATEPART(hour, CreationTime),m.line
          order by m.mes_machine,FORMAT(CreationTime, 'yyyy-MM-dd'),DATEPART(hour, CreationTime),m.line
        """
        detail_raws = db.select_sql_dict(sql)
        df_detail = pd.DataFrame(detail_raws)

        # 點數機合并主数据
        df_final = pd.merge(df_main, df_detail, on=['Name', 'Period', 'Line'], how='left')

        # 光檢機
        sql = f"""
        SELECT CAST(DATEPART(hour, Cdt) as INT) Period,m.mes_machine Name,m.line Line,Sum(OKQty) OKQty, Sum(NGQty) NGQty,
        CASE 
                WHEN Sum(NGQty) = 0 THEN 0
                ELSE Round(CAST(Sum(NGQty) AS FLOAT) / (Sum(OKQty) + Sum(NGQty)),3)
            END AS Ng_Ratio
          FROM [PMG_DEVICE].[dbo].[OpticalDevice] d, [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m
          where Cdt between CONVERT(DATETIME, '{report_date1} 05:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120)
          and d.DeviceId = m.Counting_machine
          group by m.mes_machine, DATEPART(hour, Cdt),m.line
        """
        optical_raws = db.select_sql_dict(sql)
        df_optical = pd.DataFrame(optical_raws)

        # 光檢機合并主数据
        df_final = pd.merge(df_final, df_optical, on=['Name', 'Period', 'Line'], how='left')


        # 选择需要的列 'A' 和 'B'
        df_selected = df_final[['Date', 'Name', 'Line', 'Shift', 'WorkOrderId', 'PartNo', 'ProductItem', 'AQL', 'Period', 'max_speed', 'min_speed', 'avg_speed', 'sum_qty', 'Ng_Ratio']]

        # 将数据按 'Name' 分组，并计算小计
        grouped = df_selected.groupby(['Name'])


        # 创建一个新的 DataFrame 用于保存最终的结果
        rows = []
        for name, group in grouped:
            # 按 Date 和 Period 排序
            group_sorted = group.sort_values(by=['Date', 'Shift', 'Period'])
            # 将排序后的 group 添加到 rows 列表中
            rows.append(group_sorted)

            # 合并多值的字段，用 '/' 隔开
            def join_values(col):
                return '/'.join(map(str, sorted(set(col))))

            subtotal = {
                'Date': join_values(group['Date']),
                'Name': join_values(group['Name']),
                'WorkOrderId': join_values(group['WorkOrderId']),
                'PartNo': join_values(group['PartNo']),
                'ProductItem': join_values(group['ProductItem']),
                'AQL': join_values(group['AQL']),
                'Line': 'Subtotal',
                'Shift': '',
                'Period': '',
                'max_speed': group['max_speed'].max(),
                'min_speed': group['min_speed'].min(),
                'avg_speed': group['avg_speed'].mean(),
                'sum_qty': group['sum_qty'].sum(),
                'Ng_Ratio': group['Ng_Ratio'].mean(),
            }

            # 转换为DataFrame并排序
            subtotal_df = pd.DataFrame([subtotal])
            subtotal_df = subtotal_df.sort_values(by='Date')

            rows.append(subtotal_df)

        # 将分组后的数据组合成一个 DataFrame
        df_with_subtotals = pd.concat(rows, ignore_index=True)

        # 平均車速四捨五入
        df_with_subtotals['avg_speed'] = df_final['avg_speed'].round(0)

        # 按指定顺序排列列
        column_order = [
            'Date', 'Name', 'WorkOrderId', 'PartNo', 'ProductItem', 'AQL',
            'Line', 'Shift', 'Period', 'max_speed', 'min_speed', 'avg_speed',
            'sum_qty', 'Ng_Ratio'
        ]
        df_with_subtotals = df_with_subtotals[column_order]

        # 更改列名：将 'Name' 改为 '机台号'
        df_with_subtotals.rename(columns={'Date': '生產日期'}, inplace=True)
        df_with_subtotals.rename(columns={'Name': '機台號'}, inplace=True)
        df_with_subtotals.rename(columns={'WorkOrderId': '工單'}, inplace=True)
        df_with_subtotals.rename(columns={'PartNo': '料號'}, inplace=True)
        df_with_subtotals.rename(columns={'ProductItem': '品項'}, inplace=True)
        df_with_subtotals.rename(columns={'Line': '線別'}, inplace=True)
        df_with_subtotals.rename(columns={'Shift': '班別'}, inplace=True)
        df_with_subtotals.rename(columns={'Period': '時間區段'}, inplace=True)
        df_with_subtotals.rename(columns={'max_speed': '車速(最高)'}, inplace=True)
        df_with_subtotals.rename(columns={'min_speed': '車速(最低)'}, inplace=True)
        df_with_subtotals.rename(columns={'avg_speed': '車速(平均)'}, inplace=True)
        df_with_subtotals.rename(columns={'sum_qty': '產量(加總)'}, inplace=True)
        df_with_subtotals.rename(columns={'Ng_Ratio': '光檢不良率'}, inplace=True)




        # 创建加粗字体样式
        bold_font = Font(bold=True)

        # 创建上方加粗线条的边框样式
        thick_border = Border(top=Side(style='thick'), bottom=Side(style='thick'))

        # 创建 Excel 文件
        excel_file = 'output_with_subtotals.xlsx'
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df_with_subtotals.to_excel(writer, sheet_name='Data', index=False)

            # 读取已写入的Excel文件
            workbook = writer.book
            worksheet = writer.sheets['Data']

            # 冻结第一行
            worksheet.freeze_panes = worksheet['A2']

            # 定义样式
            percent_style = NamedStyle(name='percent_style', number_format='0.00%')
            right_align_style = NamedStyle(name='right_align_style', alignment=Alignment(horizontal='right'))
            center_align_style = NamedStyle(name='center_align_style', alignment=Alignment(horizontal='center'))

            # 定义Header样式
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center')
            header_border = Border(bottom=Side(style='thin'))

            # 应用Header样式
            for cell in worksheet[1]:  # 第一行，即Header
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border

            # 设置百分比格式
            for cell in worksheet['N'][1:]:  # 'N' 是 'Ng_Ratio' 列的字母列标
                cell.style = percent_style

            # 添加分组（在 Excel 中实现点击展开/收起）
            current_row = 2
            for i, (_, group) in enumerate(grouped):
                group_size = len(group)
                start_row = current_row
                end_row = start_row + group_size - 1

                # 隐藏详细数据
                worksheet.row_dimensions.group(start_row, end_row, hidden=True)

                # 移到下一组的位置
                current_row = end_row + 2  # +2 to account for the subtotal row

            # 格式设置
            for col in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                col_letter = col[0].column_letter

                if col_letter in ['G', 'H', 'I']:  # 检查是否为指定的列
                    worksheet.column_dimensions[col_letter].width = 10  # 如果内容为空，设置为20
                else:
                    worksheet.column_dimensions[col_letter].width = max_length + 5

                # 设置对齐方式
                for cell in col:
                    if col_letter in ['J', 'K', 'L', 'M', 'N']:  # 针对指定列应用右对齐
                        cell.alignment = right_align_style.alignment
                    else:
                        cell.alignment = center_align_style.alignment

            # 遍历所有行，找到小计行并应用灰色填充、加粗字体和上方加粗线
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                if row[6].value == 'Subtotal':  # 假设“Subtotal”在第7列 ('G')
                    for cell in row:
                        cell.font = bold_font
                        cell.border = thick_border

            # 保存文件
            workbook.save(excel_file)



    def shift(self, period):
        if 6 <= int(period) <= 17:
            return '早班'
        else:
            return '晚班'



from datetime import datetime, timedelta
report_date1 = datetime.today() - timedelta(days=1)
report_date1 = report_date1.strftime('%Y%m%d')

report_date2 = datetime.today()
report_date2 = report_date2.strftime('%Y%m%d')

report = mes_daily_report(report_date1, report_date2)
report.generate_excel()