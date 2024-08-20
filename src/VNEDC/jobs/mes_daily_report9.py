import sys
import os
curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.mime.image import MIMEImage
import pandas as pd
from jobs.database import mes_database
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side, PatternFill


class mes_daily_report(object):
    report_date1 = ""
    report_date2 = ""

    # Define Style
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    right_align_style = NamedStyle(name='right_align_style', alignment=Alignment(horizontal='right'))
    center_align_style = NamedStyle(name='center_align_style', alignment=Alignment(horizontal='center'))

    # Define Header
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    header_border = Border(bottom=Side(style='thin'))

    def __init__(self, report_date1, report_date2):
        self.report_date1 = report_date1
        self.report_date2 = report_date2


    def send_email(self, file_list, data_date):
        # SMTP Sever config setting
        smtp_server = 'mail.egvnco.com'
        smtp_port = 587
        # smtp_user = 'vn_report@egvnco.com'
        # smtp_password = ''
        smtp_user = 'box.chang@egvnco.com'
        smtp_password = '1qazxsw2'

        # Receiver
        to_emails = ['box.chang@egvnco.com', 'phil.wang@egvnco.com']

        # Mail Info
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = ', '.join(to_emails)
        msg['Subject'] = f'江田廠產量日報表 {data_date}'

        # Mail Content
        html = """\
        <html>
          <body>
          </body>
        </html>
        """

        msg.attach(MIMEText(html, 'html'))

        # Attach Excel
        for file in file_list:
            excel_file = file['excel_file']
            file_name = file['file_name']
            with open(excel_file, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f"attachment; filename= {file_name}")
                msg.attach(part)

        # Send Email
        try:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, to_emails, msg.as_string())
            server.quit()
            print("Sent Email Successfully")
        except Exception as e:
            print(f"Sent Email Fail: {e}")
        finally:
            attachment.close()

    def main(self):
        report_date1 = self.report_date1
        report_date2 = self.report_date2
        db = mes_database()

        file_list = []

        # Save Path media/daily_output/
        save_path = os.path.join("..", "media", "daily_output")

        # Check folder to create
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        for plant in ['NBR', 'PVC']:

            # Create Excel file
            file_name = f'MES_OUTPUT_DAILY_Report_{report_date1}_{plant}.xlsx'
            excel_file = os.path.join(save_path, file_name)

            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:

                df_main = self.get_df_main(db, report_date1, report_date2, plant)

                df_detail = self.get_df_detail(db, report_date1, report_date2, plant)

                df_final = pd.merge(df_main, df_detail, on=['Name', 'Period', 'Line'], how='left')

                df_final['Period'] = df_final['Period'].astype(str).str.zfill(2) + ":00"

                df_selected = df_final[['Date', 'Name', 'Line', 'Shift', 'WorkOrderId', 'PartNo', 'ProductItem', 'AQL', 'Period', 'max_speed', 'min_speed', 'avg_speed', 'sum_qty']]

                df_with_subtotals = self.sorting_data(df_selected)
                self.generate_summary_excel(writer, df_with_subtotals)

                machine_groups = df_selected.groupby('Name')

                for machine_name, machine_df in machine_groups:
                    machine_df = machine_df.sort_values(by=['Date', 'Shift', 'Period'])
                    self.generate_excel(writer, machine_df, plant, machine_name)

            file_list.append({'file_name': file_name, 'excel_file': excel_file})

        # Send Email
        self.send_email(file_list, report_date1)


    def shift(self, period):
        if 6 <= int(period) <= 17:
            return '早班'
        else:
            return '晚班'

       # Work Order
    def get_df_main(self, db, report_date1, report_date2, plant):
        sql = f"""
                          SELECT w.MachineId,Name,wi.LineId Line,CAST(r.Period as INT) Period,wi.StartDate, wi.EndDate, wi.WorkOrderId,WorkOrderDate,CustomerName,PartNo,ProductItem,w.AQL,w.PlanQty,wi.Qty,w.Status
                          FROM [PMG_MES_WorkOrderInfo] wi, [PMG_MES_WorkOrder] w, [PMG_DML_DataModelList] dl,[PMG_MES_RunCard] r
                          where wi.WorkOrderId = w.id and w.MachineId = dl.Id and r.WorkOrderId = w.Id and r.LineName = wi.LineId
                          and wi.StartDate between CONVERT(DATETIME, '{report_date1} 05:30:00', 120) and CONVERT(DATETIME, '{report_date2} 05:29:59', 120)
                          and Name like '%{plant}%'
                          order by Name,CAST(r.Period as INT),wi.LineId
                        """
        raws = db.select_sql_dict(sql)

        df_main = pd.DataFrame(raws)

        # Add Column Shift
        df_main['Shift'] = df_main['Period'].apply(self.shift)

        return df_main

    # Counting Machine Data
    def get_df_detail(self, db, report_date1, report_date2, plant):

        sql = f"""
                        SELECT FORMAT(CreationTime, 'yyyy-MM-dd') AS Date,CAST(DATEPART(hour, CreationTime) as INT) Period ,m.mes_machine Name,m.line Line, max(Speed) max_speed,min(Speed) min_speed,round(avg(Speed),0) avg_speed,sum(Qty2) sum_qty
                          FROM [PMG_DEVICE].[dbo].[COUNTING_DATA] c, [PMG_DEVICE].[dbo].[COUNTING_DATA_MACHINE] m
                          where CreationTime between CONVERT(DATETIME, '{report_date1} 06:00:00', 120) and CONVERT(DATETIME, '{report_date2} 05:59:59', 120)
                          and c.MachineName = m.counting_machine and m.mes_machine like '%{plant}%'
                          group by m.mes_machine,FORMAT(CreationTime, 'yyyy-MM-dd'),DATEPART(hour, CreationTime),m.line
                          order by m.mes_machine,FORMAT(CreationTime, 'yyyy-MM-dd'),DATEPART(hour, CreationTime),m.line
                        """
        detail_raws = db.select_sql_dict(sql)
        df_detail = pd.DataFrame(detail_raws)

        return df_detail

    def generate_excel(self, writer, df, plant, machine_name):
        namesheet = str(machine_name).split('_')[-1]
        # Write data to the Excel sheet with the machine name as the sheet name
        df.to_excel(writer, sheet_name=namesheet, index=False)

        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[namesheet]

        # Freeze the first row
        worksheet.freeze_panes = worksheet['A2']

        # Apply Header Style
        for cell in worksheet[1]:  # First line is Header
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.header_border

        # Formatting
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter

            worksheet.column_dimensions[col_letter].width = max_length + 5

            # Set alignment
            for cell in col:
                if col_letter in ['J', 'K', 'L', 'M', 'N']:  # Apply right alignment for specific columns
                    cell.alignment = self.right_align_style.alignment
                else:
                    cell.alignment = self.center_align_style.alignment

        return workbook

    def generate_summary_excel(self, writer, df):
        # Create a bold font style
        bold_font = Font(bold=True)

        # Create a border style with a bold line above
        thick_border = Border(top=Side(style='thick'), bottom=Side(style='thick'))

        namesheet = "Summary"
        # Write data to the Excel sheet with the machine name as the sheet name
        df.to_excel(writer, sheet_name=namesheet, index=False)

        # Read the written Excel file
        workbook = writer.book
        worksheet = writer.sheets[namesheet]

        # Freeze the first row
        worksheet.freeze_panes = worksheet['A2']

        # Formatting
        for col in worksheet.columns:
            col_letter = col[0].column_letter
            max_length = max(len(str(cell.value)) for cell in col)

            for cell in col:
                if col_letter in ['A', 'B']:  # 检查是否为指定的列
                    worksheet.column_dimensions[col_letter].width = max_length + 5

                if col_letter in ['F', 'G', 'H']:  # 检查是否为指定的列
                    worksheet.column_dimensions[col_letter].width = 20
                    cell.alignment = self.right_align_style.alignment
                elif col_letter in ['I']:
                    worksheet.column_dimensions[col_letter].width = 20
                    cell.alignment = self.right_align_style.alignment
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = self.center_align_style.alignment


        # Search all lines, bold font and bold line above
        index_start = 2
        index_end = 2
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            if row[4].value != '':  # Line
                for cell in row[4:]:
                    cell.fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")


            if row[3].value != '':  # Shift
                # Line
                worksheet.row_dimensions.group(index_start, index_end-1, hidden=True, outline_level=2)

                worksheet.row_dimensions.group(index_end, index_end, hidden=True, outline_level=1)
                index_start = index_end + 1

                for cell in row[3:]:
                    cell.font = bold_font
                    cell.border = Border(top=Side(style='thin'))

            elif row[0].value != '':  # Machine
                # Hide detailed data
                index_start = index_end + 1
                worksheet.row_dimensions.group(index_start, index_end, hidden=True, outline_level=0)

                for cell in row:
                    cell.font = bold_font
                    cell.border = thick_border

            index_end += 1

        return workbook

    # Sorting data
    def sorting_data(self, df):
        def join_values(col):
            return '/'.join(map(str, sorted(set(col))))

        try:
            # Drop the 'Period' and 'Date' column from each group
            group_without_period = df.drop(columns=['Period', 'Date'])

            # Data group by 'Name and then calculating
            mach_grouped = group_without_period.groupby(['Name'])

            rows = []

            for mach_name, mach_group in mach_grouped:

                line_grouped = mach_group.groupby(['Shift', 'Line'])

                tmp_rows = []
                for line_name, line_group in line_grouped:
                    line_sum_df = line_group.groupby(['Name', 'ProductItem', 'Shift', 'Line']).agg({
                        'min_speed': 'min',  # Min speed
                        'max_speed': 'max',  # Max speed
                        'avg_speed': 'mean',  # Average speed
                        'sum_qty': 'sum',
                    }).reset_index()

                    # Add AQL Column to fit format
                    line_sum_df.insert(line_sum_df.columns.get_loc('ProductItem') + 1, 'AQL', None)

                    tmp_rows.append(line_sum_df)

                df_tmp = pd.concat(tmp_rows, ignore_index=True)

                # Sorting Data
                # Day Shift
                day_df = df_tmp[df_tmp['Shift'] == '早班'].copy()
                subtotal = {
                    'Name': '',
                    'ProductItem': '',
                    'AQL': '',
                    'Shift': join_values(day_df['Shift']),
                    'Line': '',
                    'max_speed': day_df['max_speed'].max(),
                    'min_speed': day_df['min_speed'].min(),
                    'avg_speed': day_df['avg_speed'].mean(),
                    'sum_qty': day_df['sum_qty'].sum(),
                }
                subtotal_df = pd.DataFrame([subtotal])
                # Average speed rounded
                subtotal_df['avg_speed'] = subtotal_df['avg_speed'].round(0)

                day_df[['Name', 'ProductItem', 'AQL', 'Shift']] = ''
                day_df['avg_speed'] = day_df['avg_speed'].round(0)
                rows.append(day_df)  # Day row data
                rows.append(subtotal_df)  # Day Shift total summary

                # Night Shift
                night_df = df_tmp[df_tmp['Shift'] == '晚班'].copy()
                subtotal = {
                    'Name': '',
                    'ProductItem': '',
                    'AQL': '',
                    'Shift': join_values(night_df['Shift']),
                    'Line': '',
                    'max_speed': night_df['max_speed'].max(),
                    'min_speed': night_df['min_speed'].min(),
                    'avg_speed': night_df['avg_speed'].mean(),
                    'sum_qty': night_df['sum_qty'].sum(),
                }
                subtotal_df = pd.DataFrame([subtotal])
                # Average speed rounded
                subtotal_df['avg_speed'] = subtotal_df['avg_speed'].round(0)

                night_df[['Name', 'ProductItem', 'AQL', 'Shift']] = ''
                night_df['avg_speed'] = night_df['avg_speed'].round(0)
                rows.append(night_df)  # Night row data
                rows.append(subtotal_df)  # Night Shift total summary

                subtotal = {
                    'Name': join_values(mach_group['Name']),
                    'ProductItem': join_values(mach_group['ProductItem']),
                    'AQL': join_values(mach_group['AQL']),
                    'Shift': '',
                    'Line': '',
                    'max_speed': mach_group['max_speed'].max(),
                    'min_speed': mach_group['min_speed'].min(),
                    'avg_speed': mach_group['avg_speed'].mean(),
                    'sum_qty': mach_group['sum_qty'].sum(),
                }
                subtotal_df = pd.DataFrame([subtotal])
                # Average speed rounded
                subtotal_df['avg_speed'] = subtotal_df['avg_speed'].round(0)
                rows.append(subtotal_df)  # Machine total summary

            # Combine the grouped data into a DataFrame
            df_with_subtotals = pd.concat(rows, ignore_index=True)

            # Change column name
            df_with_subtotals.rename(columns={'Name': '機台號'}, inplace=True)
            df_with_subtotals.rename(columns={'ProductItem': '品項'}, inplace=True)
            df_with_subtotals.rename(columns={'Line': '線別'}, inplace=True)
            df_with_subtotals.rename(columns={'Shift': '班別'}, inplace=True)
            df_with_subtotals.rename(columns={'max_speed': '車速(最高)'}, inplace=True)
            df_with_subtotals.rename(columns={'min_speed': '車速(最低)'}, inplace=True)
            df_with_subtotals.rename(columns={'avg_speed': '車速(平均)'}, inplace=True)
            df_with_subtotals.rename(columns={'sum_qty': '產量(加總)'}, inplace=True)

        except Exception as e:
            print(e)

        return df_with_subtotals

from datetime import datetime, timedelta
report_date1 = datetime.today() - timedelta(days=1)
report_date1 = report_date1.strftime('%Y%m%d')

report_date2 = datetime.today()
report_date2 = report_date2.strftime('%Y%m%d')

report = mes_daily_report(report_date1, report_date2)
report.main()
